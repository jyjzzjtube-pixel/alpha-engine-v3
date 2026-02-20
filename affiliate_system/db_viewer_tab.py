# -*- coding: utf-8 -*-
"""
DB 뷰어 탭 — API 사용량 데이터베이스 전용 조회/분석/추출 모듈
MCN 자동화 시스템 (PyQt6 다크 테마)

사용:
    from affiliate_system.db_viewer_tab import DBViewerTab
    tab = DBViewerTab()
"""
import sys
import sqlite3
import os
from pathlib import Path
from datetime import datetime, date

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QLineEdit,
    QTableWidget, QTableWidgetItem, QHeaderView, QFrame, QGroupBox,
    QGridLayout, QComboBox, QDoubleSpinBox, QSpinBox, QCheckBox,
    QDateEdit, QFileDialog, QMessageBox, QSizePolicy, QAbstractItemView,
)
from PyQt6.QtCore import Qt, QDate, QTimer
from PyQt6.QtGui import QColor, QFont

from affiliate_system.config import COST_TRACKER_DB, PROJECT_DIR

# CostTracker — 환율 조회용
sys.path.insert(0, str(PROJECT_DIR))
from api_cost_tracker import CostTracker

# Excel 지원
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font as XlFont, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─── 상수 ───
ROWS_PER_PAGE = 100
DEBOUNCE_MS = 300

# ─── 색상 팔레트 ───
CLR_BG = "#0a0e1a"
CLR_CARD = "#111827"
CLR_BORDER = "#1f2937"
CLR_ACCENT = "#6366f1"
CLR_TEXT = "#e2e8f0"
CLR_MUTED = "#6b7280"
CLR_ROW_EVEN = "#0d1224"
CLR_ROW_ODD = "#111827"
CLR_HOVER = "#1e2740"
CLR_GREEN = "#10b981"
CLR_RED = "#ef4444"
CLR_YELLOW = "#f59e0b"


def _frame_style(extra: str = "") -> str:
    return (
        f"background: {CLR_CARD}; border: 1px solid {CLR_BORDER}; "
        f"border-radius: 8px; padding: 10px; {extra}"
    )


def _btn_style(bg: str = CLR_ACCENT, hover: str = "#4f46e5") -> str:
    return (
        f"QPushButton {{ background: {bg}; color: #ffffff; border: none; "
        f"border-radius: 6px; padding: 7px 16px; font-weight: 700; font-size: 12px; }}"
        f"QPushButton:hover {{ background: {hover}; }}"
        f"QPushButton:disabled {{ background: #374151; color: #6b7280; }}"
    )


def _input_style() -> str:
    return (
        f"QLineEdit, QDoubleSpinBox, QSpinBox, QDateEdit, QComboBox {{"
        f"  background: {CLR_BG}; color: {CLR_TEXT}; border: 1px solid {CLR_BORDER};"
        f"  border-radius: 5px; padding: 5px 8px; font-size: 12px;"
        f"}}"
        f"QLineEdit:focus, QDoubleSpinBox:focus, QSpinBox:focus, QDateEdit:focus {{"
        f"  border-color: {CLR_ACCENT};"
        f"}}"
        f"QComboBox::drop-down {{ border: none; }}"
        f"QComboBox QAbstractItemView {{"
        f"  background: {CLR_CARD}; color: {CLR_TEXT}; selection-background-color: {CLR_ACCENT};"
        f"}}"
    )


class DBViewerTab(QWidget):
    """API 사용량 DB 뷰어 위젯 — 검색 / 필터 / 페이징 / 추출"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._all_rows: list[tuple] = []
        self._filtered_rows: list[tuple] = []
        self._current_page = 0
        self._total_pages = 1
        self._sort_col: int | None = None
        self._sort_asc = True

        # 환율
        self._tracker = CostTracker(db_path=COST_TRACKER_DB, project_name="db_viewer")
        self._exchange_rate = self._tracker.get_exchange_rate()  # USD->KRW
        self._jpy_rate = self._exchange_rate / 10.5  # 대략적 KRW->JPY 변환 (1 JPY ≈ 10.5 KRW)

        # 통화 표시 설정
        self._show_usd = True
        self._show_krw = True
        self._show_jpy = False

        # debounce 타이머
        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(DEBOUNCE_MS)
        self._search_timer.timeout.connect(self._apply_all_filters)

        self._build_ui()
        self._load_data()

    # ================================================================
    #  UI 구성
    # ================================================================
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 12, 16, 12)
        root.setSpacing(10)

        root.addLayout(self._build_search_bar())
        root.addWidget(self._build_filter_panel())
        root.addLayout(self._build_summary_bar())
        root.addWidget(self._build_table(), stretch=1)
        root.addLayout(self._build_pagination())

        self.setStyleSheet(_input_style())

    # ── 검색 바 ──
    def _build_search_bar(self) -> QHBoxLayout:
        row = QHBoxLayout()
        lbl = QLabel("검색")
        lbl.setStyleSheet(f"color: {CLR_TEXT}; font-weight: 700; font-size: 13px;")
        row.addWidget(lbl)

        self._search_input = QLineEdit()
        self._search_input.setPlaceholderText("ID, 프로젝트, 모델, 날짜 등 검색...")
        self._search_input.setMinimumWidth(320)
        self._search_input.textChanged.connect(self._on_search_text_changed)
        row.addWidget(self._search_input, stretch=1)

        btn_refresh = QPushButton("새로고침")
        btn_refresh.setStyleSheet(_btn_style("#374151", "#4b5563"))
        btn_refresh.clicked.connect(self._load_data)
        row.addWidget(btn_refresh)
        return row

    # ── 상세 필터 패널 ──
    def _build_filter_panel(self) -> QGroupBox:
        box = QGroupBox("상세 필터")
        box.setStyleSheet(
            f"QGroupBox {{ color: {CLR_TEXT}; border: 1px solid {CLR_BORDER}; "
            f"border-radius: 8px; margin-top: 8px; padding-top: 18px; font-weight:700; }}"
            f"QGroupBox::title {{ subcontrol-origin: margin; left: 12px; padding: 0 6px; }}"
        )
        grid = QGridLayout(box)
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(8)

        # 행 0 — 기간
        grid.addWidget(self._lbl("기간:"), 0, 0)
        self._date_from = QDateEdit()
        self._date_from.setCalendarPopup(True)
        self._date_from.setDate(QDate.currentDate().addMonths(-1))
        self._date_from.setDisplayFormat("yyyy-MM-dd")
        grid.addWidget(self._date_from, 0, 1)

        grid.addWidget(self._lbl("~"), 0, 2)
        self._date_to = QDateEdit()
        self._date_to.setCalendarPopup(True)
        self._date_to.setDate(QDate.currentDate())
        self._date_to.setDisplayFormat("yyyy-MM-dd")
        grid.addWidget(self._date_to, 0, 3)

        # 모델
        grid.addWidget(self._lbl("모델:"), 0, 4)
        self._combo_model = QComboBox()
        self._combo_model.addItem("전체")
        self._combo_model.setMinimumWidth(170)
        grid.addWidget(self._combo_model, 0, 5)

        # 비용
        grid.addWidget(self._lbl("비용(USD):"), 0, 6)
        self._cost_min = QDoubleSpinBox()
        self._cost_min.setRange(0, 9999)
        self._cost_min.setDecimals(6)
        self._cost_min.setPrefix("$")
        self._cost_min.setValue(0)
        grid.addWidget(self._cost_min, 0, 7)

        grid.addWidget(self._lbl("~"), 0, 8)
        self._cost_max = QDoubleSpinBox()
        self._cost_max.setRange(0, 9999)
        self._cost_max.setDecimals(6)
        self._cost_max.setPrefix("$")
        self._cost_max.setValue(9999)
        grid.addWidget(self._cost_max, 0, 9)

        # 행 1 — 프로젝트
        grid.addWidget(self._lbl("프로젝트:"), 1, 0)
        self._combo_project = QComboBox()
        self._combo_project.addItem("전체")
        self._combo_project.setMinimumWidth(140)
        grid.addWidget(self._combo_project, 1, 1)

        # 토큰
        grid.addWidget(self._lbl("토큰:"), 1, 2)
        self._token_min = QSpinBox()
        self._token_min.setRange(0, 99_999_999)
        self._token_min.setValue(0)
        grid.addWidget(self._token_min, 1, 3)

        grid.addWidget(self._lbl("~"), 1, 4)
        self._token_max = QSpinBox()
        self._token_max.setRange(0, 99_999_999)
        self._token_max.setValue(99_999_999)
        grid.addWidget(self._token_max, 1, 5)

        # 버튼
        btn_apply = QPushButton("필터 적용")
        btn_apply.setStyleSheet(_btn_style())
        btn_apply.clicked.connect(self._apply_all_filters)
        grid.addWidget(btn_apply, 1, 7)

        btn_reset = QPushButton("초기화")
        btn_reset.setStyleSheet(_btn_style("#374151", "#4b5563"))
        btn_reset.clicked.connect(self._reset_filters)
        grid.addWidget(btn_reset, 1, 9)

        return box

    # ── 요약 + 액션 바 ──
    def _build_summary_bar(self) -> QHBoxLayout:
        row = QHBoxLayout()

        self._lbl_total = QLabel("총 0건")
        self._lbl_total.setStyleSheet(f"color: {CLR_TEXT}; font-weight: 700; font-size: 13px;")
        row.addWidget(self._lbl_total)

        self._lbl_selected = QLabel("| 선택 0건")
        self._lbl_selected.setStyleSheet(f"color: {CLR_MUTED}; font-size: 12px;")
        row.addWidget(self._lbl_selected)

        self._lbl_cost = QLabel("| 총비용 $0.0000")
        self._lbl_cost.setStyleSheet(f"color: {CLR_GREEN}; font-weight: 700; font-size: 12px;")
        row.addWidget(self._lbl_cost)

        self._lbl_selected_cost = QLabel("")
        self._lbl_selected_cost.setStyleSheet(f"color: {CLR_YELLOW}; font-size: 12px;")
        row.addWidget(self._lbl_selected_cost)

        row.addStretch()

        # 통화 표시 체크박스
        curr_lbl = QLabel("표시:")
        curr_lbl.setStyleSheet(f"color: {CLR_MUTED}; font-size: 11px; font-weight: 700;")
        row.addWidget(curr_lbl)

        self._chk_usd = QCheckBox("$ USD")
        self._chk_usd.setChecked(True)
        self._chk_usd.setStyleSheet(f"color: {CLR_GREEN}; font-size: 11px; font-weight: 700;")
        self._chk_usd.stateChanged.connect(self._on_currency_changed)
        row.addWidget(self._chk_usd)

        self._chk_krw = QCheckBox("₩ KRW")
        self._chk_krw.setChecked(True)
        self._chk_krw.setStyleSheet(f"color: {CLR_YELLOW}; font-size: 11px; font-weight: 700;")
        self._chk_krw.stateChanged.connect(self._on_currency_changed)
        row.addWidget(self._chk_krw)

        self._chk_jpy = QCheckBox("¥ JPY")
        self._chk_jpy.setChecked(False)
        self._chk_jpy.setStyleSheet(f"color: #06b6d4; font-size: 11px; font-weight: 700;")
        self._chk_jpy.stateChanged.connect(self._on_currency_changed)
        row.addWidget(self._chk_jpy)

        row.addWidget(self._separator_v())

        # 전체 선택 / 해제
        btn_sel_all = QPushButton("전체 선택")
        btn_sel_all.setStyleSheet(_btn_style("#374151", "#4b5563"))
        btn_sel_all.clicked.connect(lambda: self._toggle_all_checks(True))
        row.addWidget(btn_sel_all)

        btn_desel = QPushButton("전체 해제")
        btn_desel.setStyleSheet(_btn_style("#374151", "#4b5563"))
        btn_desel.clicked.connect(lambda: self._toggle_all_checks(False))
        row.addWidget(btn_desel)

        row.addWidget(self._separator_v())

        # 엑셀 추출
        btn_xls_sel = QPushButton("엑셀 추출 (선택)")
        btn_xls_sel.setStyleSheet(_btn_style("#065f46", "#047857"))
        btn_xls_sel.clicked.connect(lambda: self._export_excel(selected_only=True))
        row.addWidget(btn_xls_sel)

        btn_xls_all = QPushButton("엑셀 추출 (전체)")
        btn_xls_all.setStyleSheet(_btn_style("#065f46", "#047857"))
        btn_xls_all.clicked.connect(lambda: self._export_excel(selected_only=False))
        row.addWidget(btn_xls_all)

        btn_pdf = QPushButton("PDF 추출 (선택)")
        btn_pdf.setStyleSheet(_btn_style("#7c2d12", "#9a3412"))
        btn_pdf.clicked.connect(self._export_pdf_selected)
        row.addWidget(btn_pdf)

        row.addWidget(self._separator_v())

        btn_del = QPushButton("선택 삭제")
        btn_del.setStyleSheet(_btn_style(CLR_RED, "#dc2626"))
        btn_del.clicked.connect(self._on_bulk_delete)
        row.addWidget(btn_del)

        return row

    # ── 데이터 테이블 ──
    def _build_table(self) -> QTableWidget:
        cols = ["", "ID", "시각", "프로젝트", "모델",
                "Input 토큰", "Output 토큰", "비용(USD)", "비용(KRW)"]
        # 통화 체크박스가 아직 생성 안됐을 수 있으므로 기본값 사용
        self._table = QTableWidget(0, len(cols))
        self._table.setHorizontalHeaderLabels(cols)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.verticalHeader().setVisible(False)
        self._table.setAlternatingRowColors(False)  # 직접 처리
        self._table.setSortingEnabled(False)
        self._table.horizontalHeader().sectionClicked.connect(self._on_header_clicked)
        self._table.setShowGrid(False)

        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self._table.setColumnWidth(0, 36)
        for c in range(1, len(cols)):
            hdr.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)

        self._table.setStyleSheet(f"""
            QTableWidget {{
                background: {CLR_CARD};
                border: 1px solid {CLR_BORDER};
                border-radius: 8px;
                gridline-color: {CLR_BORDER};
                color: {CLR_TEXT};
                font-size: 12px;
            }}
            QTableWidget::item {{
                padding: 4px 8px;
                border-bottom: 1px solid {CLR_BORDER};
            }}
            QTableWidget::item:selected {{
                background: {CLR_HOVER};
            }}
            QHeaderView::section {{
                background: #0d1224;
                color: {CLR_MUTED};
                font-weight: 700;
                font-size: 11px;
                border: none;
                border-bottom: 2px solid {CLR_BORDER};
                padding: 6px 4px;
            }}
        """)
        return self._table

    # ── 페이지네이션 ──
    def _build_pagination(self) -> QHBoxLayout:
        row = QHBoxLayout()
        row.addStretch()

        self._btn_prev = QPushButton("이전")
        self._btn_prev.setStyleSheet(_btn_style("#374151", "#4b5563"))
        self._btn_prev.clicked.connect(self._page_prev)
        row.addWidget(self._btn_prev)

        self._page_label = QLabel("1 / 1")
        self._page_label.setStyleSheet(
            f"color: {CLR_TEXT}; font-weight: 700; padding: 0 16px; font-size: 13px;"
        )
        row.addWidget(self._page_label)

        self._btn_next = QPushButton("다음")
        self._btn_next.setStyleSheet(_btn_style("#374151", "#4b5563"))
        self._btn_next.clicked.connect(self._page_next)
        row.addWidget(self._btn_next)

        row.addStretch()
        return row

    # ── 유틸 위젯 ──
    @staticmethod
    def _lbl(text: str) -> QLabel:
        lb = QLabel(text)
        lb.setStyleSheet(f"color: {CLR_MUTED}; font-size: 12px;")
        return lb

    @staticmethod
    def _separator_v() -> QFrame:
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.VLine)
        sep.setStyleSheet(f"color: {CLR_BORDER};")
        return sep

    # ================================================================
    #  데이터 로드
    # ================================================================
    def _load_data(self):
        """DB에서 전체 데이터를 읽고, 콤보박스 갱신 후 필터 적용."""
        db_path = COST_TRACKER_DB
        if not os.path.exists(db_path):
            self._all_rows = []
            self._apply_all_filters()
            return

        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        try:
            cur.execute(
                "SELECT id, timestamp, project, model, "
                "input_tokens, output_tokens, cost_usd FROM api_usage "
                "ORDER BY id DESC"
            )
            self._all_rows = cur.fetchall()
        except sqlite3.OperationalError:
            self._all_rows = []
        conn.close()

        self._populate_combos()
        self._apply_all_filters()

    def _populate_combos(self):
        """프로젝트 / 모델 콤보박스를 DB 값으로 채움."""
        models = sorted({r[3] for r in self._all_rows})
        projects = sorted({r[2] for r in self._all_rows})

        cur_model = self._combo_model.currentText()
        self._combo_model.blockSignals(True)
        self._combo_model.clear()
        self._combo_model.addItem("전체")
        self._combo_model.addItems(models)
        idx = self._combo_model.findText(cur_model)
        if idx >= 0:
            self._combo_model.setCurrentIndex(idx)
        self._combo_model.blockSignals(False)

        cur_proj = self._combo_project.currentText()
        self._combo_project.blockSignals(True)
        self._combo_project.clear()
        self._combo_project.addItem("전체")
        self._combo_project.addItems(projects)
        idx = self._combo_project.findText(cur_proj)
        if idx >= 0:
            self._combo_project.setCurrentIndex(idx)
        self._combo_project.blockSignals(False)

    # ================================================================
    #  검색 & 필터
    # ================================================================
    def _on_search_text_changed(self):
        self._search_timer.start()

    def _apply_all_filters(self):
        """텍스트 검색 + 상세 필터를 결합하여 _filtered_rows 산출."""
        search_text = self._search_input.text().strip().lower()

        date_from = self._date_from.date().toString("yyyy-MM-dd")
        date_to = self._date_to.date().toString("yyyy-MM-dd") + " 23:59:59"
        sel_model = self._combo_model.currentText()
        sel_project = self._combo_project.currentText()
        cost_lo = self._cost_min.value()
        cost_hi = self._cost_max.value()
        tok_lo = self._token_min.value()
        tok_hi = self._token_max.value()

        results = []
        for row in self._all_rows:
            rid, ts, proj, model, in_tok, out_tok, cost = row

            # 날짜
            if ts < date_from or ts > date_to:
                continue
            # 모델
            if sel_model != "전체" and model != sel_model:
                continue
            # 프로젝트
            if sel_project != "전체" and proj != sel_project:
                continue
            # 비용
            if cost < cost_lo or cost > cost_hi:
                continue
            # 토큰 (input + output 합계)
            total_tok = (in_tok or 0) + (out_tok or 0)
            if total_tok < tok_lo or total_tok > tok_hi:
                continue
            # 전문 검색
            if search_text:
                haystack = f"{rid} {ts} {proj} {model} {in_tok} {out_tok} {cost}".lower()
                if search_text not in haystack:
                    continue
            results.append(row)

        # 정렬 유지
        if self._sort_col is not None:
            col_idx = self._sort_col
            try:
                results.sort(key=lambda r: r[col_idx], reverse=not self._sort_asc)
            except (IndexError, TypeError):
                pass

        self._filtered_rows = results
        self._current_page = 0
        self._render_page()
        self._update_summary()

    def _reset_filters(self):
        self._search_input.clear()
        self._date_from.setDate(QDate.currentDate().addMonths(-1))
        self._date_to.setDate(QDate.currentDate())
        self._combo_model.setCurrentIndex(0)
        self._combo_project.setCurrentIndex(0)
        self._cost_min.setValue(0)
        self._cost_max.setValue(9999)
        self._token_min.setValue(0)
        self._token_max.setValue(99_999_999)
        self._sort_col = None
        self._sort_asc = True
        self._apply_all_filters()

    # ================================================================
    #  정렬
    # ================================================================
    def _on_currency_changed(self):
        """통화 체크박스 변경 시 테이블 컬럼 및 데이터를 재구성."""
        self._show_usd = self._chk_usd.isChecked()
        self._show_krw = self._chk_krw.isChecked()
        self._show_jpy = self._chk_jpy.isChecked()
        self._rebuild_table_columns()
        self._render_page()
        self._update_summary()

    def _rebuild_table_columns(self):
        """선택된 통화에 따라 테이블 컬럼을 재구성."""
        cols = ["", "ID", "시각", "프로젝트", "모델", "Input 토큰", "Output 토큰"]
        if self._show_usd:
            cols.append("비용(USD)")
        if self._show_krw:
            cols.append("비용(KRW)")
        if self._show_jpy:
            cols.append("비용(JPY)")
        self._table.setColumnCount(len(cols))
        self._table.setHorizontalHeaderLabels(cols)

        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self._table.setColumnWidth(0, 36)
        for c in range(1, len(cols)):
            hdr.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
        if len(cols) > 1:
            hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)

    def _on_header_clicked(self, logical_index: int):
        if logical_index == 0:
            return  # 체크박스 열
        # 열 인덱스 → row tuple 인덱스 (offset -1 for checkbox col)
        col_map = {1: 0, 2: 1, 3: 2, 4: 3, 5: 4, 6: 5, 7: 6, 8: 6}
        mapped = col_map.get(logical_index)
        if mapped is None:
            return
        if self._sort_col == mapped:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = mapped
            self._sort_asc = True

        try:
            self._filtered_rows.sort(
                key=lambda r: r[mapped] if r[mapped] is not None else 0,
                reverse=not self._sort_asc,
            )
        except (IndexError, TypeError):
            return
        self._current_page = 0
        self._render_page()

    # ================================================================
    #  페이지네이션
    # ================================================================
    def _render_page(self):
        total = len(self._filtered_rows)
        self._total_pages = max(1, (total + ROWS_PER_PAGE - 1) // ROWS_PER_PAGE)
        if self._current_page >= self._total_pages:
            self._current_page = self._total_pages - 1

        start = self._current_page * ROWS_PER_PAGE
        end = min(start + ROWS_PER_PAGE, total)
        page_rows = self._filtered_rows[start:end]

        self._table.setRowCount(0)
        self._table.setRowCount(len(page_rows))

        rate = self._exchange_rate
        for i, row in enumerate(page_rows):
            rid, ts, proj, model, in_tok, out_tok, cost = row
            bg = QColor(CLR_ROW_EVEN) if i % 2 == 0 else QColor(CLR_ROW_ODD)

            # 체크박스
            chk = QCheckBox()
            chk.setStyleSheet("QCheckBox { margin-left: 8px; }")
            chk.stateChanged.connect(self._update_selection_summary)
            chk_widget = QWidget()
            chk_layout = QHBoxLayout(chk_widget)
            chk_layout.addWidget(chk)
            chk_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            chk_layout.setContentsMargins(0, 0, 0, 0)
            self._table.setCellWidget(i, 0, chk_widget)

            # 데이터 셀 (동적 통화 컬럼)
            values = [
                str(rid),
                ts or "",
                proj or "",
                model or "",
                f"{in_tok:,}" if in_tok else "0",
                f"{out_tok:,}" if out_tok else "0",
            ]
            if self._show_usd:
                values.append(f"${cost:.6f}" if cost else "$0")
            if self._show_krw:
                values.append(f"₩{int(cost * rate):,}" if cost else "₩0")
            if self._show_jpy:
                jpy_val = cost * rate / 10.5 if cost else 0
                values.append(f"¥{int(jpy_val):,}" if cost else "¥0")
            for c, val in enumerate(values, start=1):
                item = QTableWidgetItem(val)
                item.setBackground(bg)
                item.setForeground(QColor(CLR_TEXT))
                # 숫자 열은 우측 정렬
                if c >= 5:
                    item.setTextAlignment(
                        Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
                    )
                else:
                    item.setTextAlignment(
                        Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
                    )
                self._table.setItem(i, c, item)

            self._table.setRowHeight(i, 32)

        # 페이지 라벨
        self._page_label.setText(f"{self._current_page + 1} / {self._total_pages}")
        self._btn_prev.setEnabled(self._current_page > 0)
        self._btn_next.setEnabled(self._current_page < self._total_pages - 1)

    def _page_prev(self):
        if self._current_page > 0:
            self._current_page -= 1
            self._render_page()

    def _page_next(self):
        if self._current_page < self._total_pages - 1:
            self._current_page += 1
            self._render_page()

    # ================================================================
    #  요약 통계
    # ================================================================
    def _update_summary(self):
        total_count = len(self._filtered_rows)
        total_cost = sum(r[6] for r in self._filtered_rows if r[6])
        rate = self._exchange_rate

        self._lbl_total.setText(f"총 {total_count:,}건")

        cost_parts = []
        if self._show_usd:
            cost_parts.append(f"${total_cost:.4f}")
        if self._show_krw:
            cost_parts.append(f"₩{int(total_cost * rate):,}")
        if self._show_jpy:
            cost_parts.append(f"¥{int(total_cost * rate / 10.5):,}")
        self._lbl_cost.setText(f"| 총비용 {' / '.join(cost_parts)}" if cost_parts else "| 총비용 -")
        self._update_selection_summary()

    def _update_selection_summary(self):
        checked = self._get_checked_indices()
        count = len(checked)
        if count == 0:
            self._lbl_selected.setText("| 선택 0건")
            self._lbl_selected_cost.setText("")
            return
        page_start = self._current_page * ROWS_PER_PAGE
        sel_cost = 0.0
        for idx in checked:
            abs_idx = page_start + idx
            if abs_idx < len(self._filtered_rows):
                sel_cost += self._filtered_rows[abs_idx][6] or 0
        rate = self._exchange_rate
        parts = []
        if self._show_usd:
            parts.append(f"${sel_cost:.4f}")
        if self._show_krw:
            parts.append(f"₩{int(sel_cost * rate):,}")
        if self._show_jpy:
            parts.append(f"¥{int(sel_cost * rate / 10.5):,}")
        self._lbl_selected.setText(f"| 선택 {count}건")
        self._lbl_selected_cost.setText(f"| 선택 비용 {' / '.join(parts)}" if parts else "")

    # ================================================================
    #  체크박스
    # ================================================================
    def _get_checked_indices(self) -> list[int]:
        """현재 페이지에서 체크된 행 인덱스 목록."""
        indices = []
        for i in range(self._table.rowCount()):
            widget = self._table.cellWidget(i, 0)
            if widget:
                chk = widget.findChild(QCheckBox)
                if chk and chk.isChecked():
                    indices.append(i)
        return indices

    def _get_checked_rows(self) -> list[tuple]:
        """체크된 행의 원본 데이터."""
        page_start = self._current_page * ROWS_PER_PAGE
        rows = []
        for idx in self._get_checked_indices():
            abs_idx = page_start + idx
            if abs_idx < len(self._filtered_rows):
                rows.append(self._filtered_rows[abs_idx])
        return rows

    def _toggle_all_checks(self, state: bool):
        for i in range(self._table.rowCount()):
            widget = self._table.cellWidget(i, 0)
            if widget:
                chk = widget.findChild(QCheckBox)
                if chk:
                    chk.setChecked(state)
        self._update_selection_summary()

    # ================================================================
    #  엑셀 추출
    # ================================================================
    def _export_excel(self, selected_only: bool = False):
        if not HAS_OPENPYXL:
            QMessageBox.warning(
                self, "라이브러리 없음",
                "openpyxl 패키지가 설치되어 있지 않습니다.\npip install openpyxl",
            )
            return

        if selected_only:
            rows = self._get_checked_rows()
            if not rows:
                QMessageBox.information(self, "안내", "선택된 항목이 없습니다.")
                return
        else:
            rows = self._filtered_rows

        if not rows:
            QMessageBox.information(self, "안내", "추출할 데이터가 없습니다.")
            return

        default_name = f"api_usage_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "엑셀 저장", str(PROJECT_DIR / default_name),
            "Excel Files (*.xlsx)",
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "API 사용량"

        headers = ["ID", "시각", "프로젝트", "모델",
                    "Input 토큰", "Output 토큰", "비용(USD)", "비용(KRW)"]

        # 헤더 스타일
        hdr_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
        hdr_font = XlFont(bold=True, color="E2E8F0", size=11)
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin_side = Side(style="thin", color="1F2937")
        hdr_border = Border(bottom=thin_side)

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = hdr_border

        rate = self._exchange_rate
        num_font = XlFont(size=10)
        for r_idx, row in enumerate(rows, 2):
            rid, ts, proj, model, in_tok, out_tok, cost = row
            vals = [rid, ts, proj, model, in_tok, out_tok,
                    round(cost, 6) if cost else 0,
                    int(cost * rate) if cost else 0]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r_idx, column=c, value=v)
                cell.font = num_font
                if c >= 5:
                    cell.alignment = Alignment(horizontal="right")
                    if c == 7:
                        cell.number_format = "$#,##0.000000"
                    elif c == 8:
                        cell.number_format = "#,##0"

        # 열 너비
        widths = [8, 22, 16, 28, 14, 14, 16, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        # 합계 행
        sum_row = len(rows) + 2
        ws.cell(row=sum_row, column=5, value="합계").font = XlFont(bold=True, size=10)
        ws.cell(row=sum_row, column=6, value=sum(r[4] or 0 for r in rows))
        ws.cell(row=sum_row, column=7, value="").font = XlFont(bold=True, size=10)
        total_cost = sum(r[6] or 0 for r in rows)
        ws.cell(row=sum_row, column=7, value=round(total_cost, 6))
        ws.cell(row=sum_row, column=7).number_format = "$#,##0.000000"
        ws.cell(row=sum_row, column=8, value=int(total_cost * rate))
        ws.cell(row=sum_row, column=8).number_format = "#,##0"

        try:
            wb.save(path)
            QMessageBox.information(
                self, "완료", f"엑셀 파일 저장 완료\n{path}\n({len(rows)}건)",
            )
        except Exception as e:
            QMessageBox.critical(self, "오류", f"저장 실패: {e}")

    # ================================================================
    #  PDF 추출 (HTML → 파일)
    # ================================================================
    def _export_pdf_selected(self):
        rows = self._get_checked_rows()
        if not rows:
            QMessageBox.information(self, "안내", "선택된 항목이 없습니다.")
            return
        self._export_pdf(rows)

    def _export_pdf(self, rows: list[tuple]):
        default_name = f"api_usage_{datetime.now():%Y%m%d_%H%M%S}.html"
        path, _ = QFileDialog.getSaveFileName(
            self, "PDF(HTML) 저장", str(PROJECT_DIR / default_name),
            "HTML Files (*.html);;All Files (*)",
        )
        if not path:
            return

        rate = self._exchange_rate
        total_cost = sum(r[6] or 0 for r in rows)
        total_krw = int(total_cost * rate)
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        table_rows = ""
        for i, row in enumerate(rows):
            rid, ts, proj, model, in_tok, out_tok, cost = row
            bg = "#111827" if i % 2 == 0 else "#0d1224"
            krw = int((cost or 0) * rate)
            table_rows += (
                f'<tr style="background:{bg}">'
                f"<td>{rid}</td><td>{ts}</td><td>{proj}</td><td>{model}</td>"
                f"<td style='text-align:right'>{in_tok:,}</td>"
                f"<td style='text-align:right'>{out_tok:,}</td>"
                f"<td style='text-align:right'>${cost:.6f}</td>"
                f"<td style='text-align:right'>{krw:,}</td>"
                f"</tr>\n"
            )

        html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<title>API 사용량 리포트</title>
<style>
  body {{ font-family: 'Malgun Gothic','Segoe UI',sans-serif; background:#0a0e1a;
         color:#e2e8f0; padding:24px; }}
  h1 {{ color:#f9fafb; font-size:22px; border-bottom:2px solid #6366f1; padding-bottom:8px; }}
  .meta {{ color:#6b7280; font-size:12px; margin-bottom:16px; }}
  .summary {{ background:#111827; border:1px solid #1f2937; border-radius:8px;
              padding:14px; margin-bottom:16px; }}
  .summary span {{ font-weight:700; color:#10b981; }}
  table {{ width:100%; border-collapse:collapse; font-size:12px; }}
  th {{ background:#0d1224; color:#6b7280; text-align:left; padding:8px 6px;
       border-bottom:2px solid #1f2937; font-weight:700; }}
  td {{ padding:6px; border-bottom:1px solid #1f2937; color:#e2e8f0; }}
</style>
</head>
<body>
<h1>API 사용량 리포트</h1>
<div class="meta">생성일시: {now_str} | 환율: 1 USD = {rate:,.2f} KRW</div>
<div class="summary">
  총 <span>{len(rows):,}건</span> |
  총비용 <span>${total_cost:.4f}</span>
  (<span>{total_krw:,}</span>)
</div>
<table>
<thead>
<tr><th>ID</th><th>시각</th><th>프로젝트</th><th>모델</th>
    <th>Input 토큰</th><th>Output 토큰</th><th>비용(USD)</th><th>비용(KRW)</th></tr>
</thead>
<tbody>
{table_rows}
</tbody>
<tfoot>
<tr style="background:#111827; font-weight:700;">
  <td colspan="4">합계</td>
  <td style="text-align:right">{sum(r[4] or 0 for r in rows):,}</td>
  <td style="text-align:right">{sum(r[5] or 0 for r in rows):,}</td>
  <td style="text-align:right">${total_cost:.6f}</td>
  <td style="text-align:right">{total_krw:,}</td>
</tr>
</tfoot>
</table>
</body>
</html>"""

        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(html)
            QMessageBox.information(
                self, "완료",
                f"HTML 리포트 저장 완료\n{path}\n({len(rows)}건)\n\n"
                "브라우저에서 열어 인쇄(PDF) 할 수 있습니다.",
            )
        except Exception as e:
            QMessageBox.critical(self, "오류", f"저장 실패: {e}")

    # ================================================================
    #  삭제 (관리자 전용 — 비활성)
    # ================================================================
    def _on_bulk_delete(self):
        QMessageBox.information(
            self, "관리자 전용",
            "데이터 삭제는 관리자 전용 기능입니다.\n직접 DB를 조작하세요.",
        )

    # ================================================================
    #  외부 인터페이스
    # ================================================================
    def refresh(self):
        """외부에서 탭 전환 시 호출 가능."""
        self._exchange_rate = self._tracker.get_exchange_rate()
        self._load_data()


# ─── 독립 실행 테스트 ───
if __name__ == "__main__":
    from PyQt6.QtWidgets import QApplication

    app = QApplication(sys.argv)
    app.setStyleSheet(f"""
        QWidget {{
            background-color: {CLR_BG};
            color: {CLR_TEXT};
            font-family: 'Malgun Gothic', 'Segoe UI', sans-serif;
            font-size: 13px;
        }}
    """)
    win = DBViewerTab()
    win.setWindowTitle("DB 뷰어 - API 사용량")
    win.resize(1280, 720)
    win.show()
    sys.exit(app.exec())
