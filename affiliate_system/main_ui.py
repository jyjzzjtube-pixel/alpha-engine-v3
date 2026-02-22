"""
YJ Partners MCN & F&B 자동화 파이프라인 — 커맨드센터
PyQt6 기반 프로페셔널 다크테마 대시보드

사용법:
    python -m affiliate_system.main_ui
"""
import sys
import uuid
import os
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout,
    QHBoxLayout, QGridLayout, QLabel, QPushButton, QLineEdit,
    QTextEdit, QProgressBar, QTableWidget, QTableWidgetItem,
    QFileDialog, QComboBox, QSpinBox, QDoubleSpinBox, QCheckBox,
    QGroupBox, QSplitter, QStatusBar, QMessageBox, QHeaderView,
    QFrame, QScrollArea, QStackedWidget,
)
from PyQt6.QtCore import (
    Qt, QThread, pyqtSignal, pyqtSlot, QTimer, QSize,
)
from PyQt6.QtGui import (
    QFont, QColor, QIcon, QPixmap, QDragEnterEvent, QDropEvent,
    QAction, QPainter, QLinearGradient, QPen, QBrush,
)

from api_cost_tracker import CostTracker
from affiliate_system.config import (
    BUDGET_LIMIT_KRW, COST_TRACKER_DB, GEMINI_API_KEY,
    ANTHROPIC_API_KEY, INSTAGRAM_USERNAME,
    RENDER_OUTPUT_DIR, PROJECT_DIR, WORK_DIR,
    PEXELS_API_KEY, PIXABAY_API_KEY, UNSPLASH_ACCESS_KEY,
    TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID, CDP_URL,
    COUPANG_ACCESS_KEY, COUPANG_SECRET_KEY,
)
from affiliate_system.models import (
    Campaign, CampaignStatus, Platform, Product, AIContent, RenderConfig,
)
from affiliate_system.editor_tab import EditorTab
from affiliate_system.db_viewer_tab import DBViewerTab
from affiliate_system.ai_review_tab import AIReviewTab


# ══════════════════════════════════════════════════════════════
#  다크 테마 스타일시트 (MCN Agency Executive 스타일)
# ══════════════════════════════════════════════════════════════

DARK_STYLESHEET = """
QMainWindow, QWidget {
    background-color: #0a0e1a;
    color: #e2e8f0;
    font-family: 'Malgun Gothic', 'Segoe UI', sans-serif;
    font-size: 13px;
}
QTabWidget::pane {
    border: 1px solid #1a1f35;
    background: #111827;
    border-radius: 10px;
    top: -1px;
}
QTabBar::tab {
    background: #0a0e1a;
    color: #4b5563;
    padding: 14px 28px;
    border-top-left-radius: 10px;
    border-top-right-radius: 10px;
    margin-right: 3px;
    font-weight: 700;
    font-size: 14px;
    min-width: 100px;
}
QTabBar::tab:selected {
    background: #111827;
    color: #f9fafb;
    border-bottom: 3px solid #6366f1;
}
QTabBar::tab:hover:!selected {
    color: #9ca3af;
    background: #111827;
}
QLabel {
    color: #e2e8f0;
    background: transparent;
}
QLabel#metricValue {
    font-size: 32px;
    font-weight: 900;
    color: #f9fafb;
}
QLabel#metricLabel {
    font-size: 11px;
    color: #6b7280;
    font-weight: 700;
}
QLabel#sectionTitle {
    font-size: 17px;
    font-weight: 800;
    color: #f9fafb;
    padding: 6px 0;
}
QPushButton {
    background-color: #6366f1;
    color: white;
    border: none;
    border-radius: 10px;
    padding: 11px 22px;
    font-weight: 700;
    font-size: 13px;
}
QPushButton:hover { background-color: #4f46e5; }
QPushButton:pressed { background-color: #4338ca; }
QPushButton:disabled { background-color: #1f2937; color: #4b5563; }
QPushButton#dangerBtn { background-color: #dc2626; }
QPushButton#dangerBtn:hover { background-color: #b91c1c; }
QPushButton#successBtn { background-color: #16a34a; }
QPushButton#successBtn:hover { background-color: #15803d; }
QPushButton#secondaryBtn { background-color: #1f2937; color: #e5e7eb; }
QPushButton#secondaryBtn:hover { background-color: #374151; }
QPushButton#ghostBtn {
    background-color: transparent;
    color: #6366f1;
    border: 1px solid #6366f1;
}
QPushButton#ghostBtn:hover { background-color: rgba(99, 102, 241, 0.1); }

QLineEdit, QTextEdit, QComboBox, QSpinBox, QDoubleSpinBox {
    background-color: #111827;
    color: #e5e7eb;
    border: 1px solid #1f2937;
    border-radius: 8px;
    padding: 10px 12px;
    font-size: 13px;
    selection-background-color: #6366f1;
}
QLineEdit:focus, QTextEdit:focus { border-color: #6366f1; }
QComboBox::drop-down { border: none; width: 30px; }
QComboBox QAbstractItemView {
    background-color: #111827; color: #e5e7eb;
    border: 1px solid #1f2937;
    selection-background-color: #6366f1;
}
QGroupBox {
    color: #f9fafb;
    border: 1px solid #1f2937;
    border-radius: 12px;
    margin-top: 18px;
    padding: 22px 16px 16px 16px;
    font-weight: 700;
    font-size: 14px;
    background-color: #111827;
}
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 2px 14px;
    background-color: #111827;
    border-radius: 6px;
}
QProgressBar {
    border: none; border-radius: 6px;
    background: #1f2937; height: 10px;
    text-align: center; color: transparent;
}
QProgressBar::chunk {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #6366f1, stop:1 #a855f7);
    border-radius: 6px;
}
QTableWidget {
    background-color: #0a0e1a; color: #e5e7eb;
    gridline-color: #111827;
    border: 1px solid #1f2937; border-radius: 8px;
    font-size: 12px;
}
QTableWidget::item { padding: 8px; border-bottom: 1px solid #111827; }
QTableWidget::item:selected { background-color: rgba(99, 102, 241, 0.25); }
QHeaderView::section {
    background-color: #111827; color: #9ca3af;
    padding: 10px 8px; border: none;
    border-bottom: 2px solid #1f2937;
    font-weight: 700; font-size: 11px;
}
QStatusBar {
    background: #0a0e1a; color: #6b7280;
    border-top: 1px solid #111827; padding: 4px 12px; font-size: 12px;
}
QScrollBar:vertical {
    background: transparent; width: 8px; margin: 0;
}
QScrollBar::handle:vertical {
    background: #1f2937; border-radius: 4px; min-height: 30px;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
QCheckBox { spacing: 8px; color: #e5e7eb; font-size: 13px; }
QCheckBox::indicator {
    width: 18px; height: 18px; border-radius: 5px;
    border: 2px solid #374151; background: #111827;
}
QCheckBox::indicator:checked { background: #6366f1; border-color: #6366f1; }
QScrollArea { border: none; background: transparent; }
QToolTip {
    background-color: #111827; color: #e5e7eb;
    border: 1px solid #1f2937; border-radius: 6px; padding: 8px;
}
"""


# ══════════════════════════════════════════════════════════════
#  지표 카드 위젯
# ══════════════════════════════════════════════════════════════

class MetricCard(QFrame):
    """대시보드 상단 지표 카드"""

    def __init__(self, title: str, value: str = "0",
                 subtitle: str = "", accent: str = "#6366f1"):
        super().__init__()
        self.setFixedHeight(140)
        self.setStyleSheet(f"""
            MetricCard {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #111827, stop:1 #0a0e1a);
                border: 1px solid #1f2937;
                border-radius: 14px;
                border-left: 4px solid {accent};
            }}
        """)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 14, 18, 14)
        layout.setSpacing(4)

        lbl_title = QLabel(title)
        lbl_title.setObjectName("metricLabel")
        layout.addWidget(lbl_title)

        self.lbl_value = QLabel(value)
        self.lbl_value.setObjectName("metricValue")
        layout.addWidget(self.lbl_value)

        self.lbl_subtitle = QLabel(subtitle)
        self.lbl_subtitle.setStyleSheet("color: #6b7280; font-size: 11px;")
        layout.addWidget(self.lbl_subtitle)
        layout.addStretch()

    def set_value(self, value: str, subtitle: str = ""):
        self.lbl_value.setText(value)
        if subtitle:
            self.lbl_subtitle.setText(subtitle)


# ══════════════════════════════════════════════════════════════
#  상태 LED 위젯
# ══════════════════════════════════════════════════════════════

class StatusLED(QFrame):
    """시스템 상태 LED 인디케이터"""

    def __init__(self, label: str, connected: bool = False):
        super().__init__()
        self._name = label
        self.setFixedHeight(36)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(12, 4, 12, 4)
        layout.setSpacing(8)

        self.dot = QLabel("●")
        self.dot.setFixedWidth(16)
        self.label = QLabel(label)
        self.label.setStyleSheet("font-size: 12px; font-weight: 600;")

        layout.addWidget(self.dot)
        layout.addWidget(self.label)
        layout.addStretch()

        self.set_status(connected)

    def set_status(self, connected: bool):
        color = "#22c55e" if connected else "#ef4444"
        text = "연결됨" if connected else "미연결"
        self.dot.setStyleSheet(f"color: {color}; font-size: 16px;")
        self.label.setText(f"{self._name} — {text}")


# ══════════════════════════════════════════════════════════════
#  실시간 콘솔 로그 위젯
# ══════════════════════════════════════════════════════════════

class LiveConsole(QTextEdit):
    """실시간 백그라운드 프로세스 로그"""

    def __init__(self):
        super().__init__()
        self.setReadOnly(True)
        self.setMaximumHeight(180)
        self.setStyleSheet("""
            QTextEdit {
                background: #0a0e1a;
                color: #22c55e;
                border: 1px solid #1f2937;
                border-radius: 8px;
                font-family: 'Consolas', 'D2Coding', monospace;
                font-size: 11px;
                padding: 8px;
            }
        """)
        self.log("시스템 초기화 중...")

    def log(self, msg: str):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.append(f"[{timestamp}] {msg}")
        self.verticalScrollBar().setValue(
            self.verticalScrollBar().maximum())


# ══════════════════════════════════════════════════════════════
#  상태 컬러 매핑 (한글)
# ══════════════════════════════════════════════════════════════

STATUS_COLORS = {
    CampaignStatus.DRAFT:      ("#1f2937", "#9ca3af", "대기"),
    CampaignStatus.SCRAPING:   ("#1e3a5f", "#60a5fa", "수집중"),
    CampaignStatus.GENERATING: ("#3b1f6e", "#a78bfa", "생성중"),
    CampaignStatus.RENDERING:  ("#713f12", "#fbbf24", "렌더링"),
    CampaignStatus.UPLOADING:  ("#064e3b", "#34d399", "업로드"),
    CampaignStatus.COMPLETE:   ("#14532d", "#22c55e", "완료"),
    CampaignStatus.ERROR:      ("#7f1d1d", "#ef4444", "오류"),
}


# ══════════════════════════════════════════════════════════════
#  QThread 워커들 (UI 프리징 방지)
# ══════════════════════════════════════════════════════════════

class ScrapeWorker(QThread):
    """백그라운드 상품 스크래핑"""
    progress = pyqtSignal(str, int)
    product_ready = pyqtSignal(object)
    error = pyqtSignal(str)
    finished_all = pyqtSignal(list)

    def __init__(self, urls: list):
        super().__init__()
        self.urls = urls

    def run(self):
        products = []
        for i, url in enumerate(self.urls):
            try:
                self.progress.emit(
                    f"스크래핑 중... {i+1}/{len(self.urls)}",
                    int((i / len(self.urls)) * 100))
                product = Product(
                    url=url,
                    title=f"상품 {i+1} (스크래핑 완료)",
                    scraped_at=datetime.now())
                products.append(product)
                self.product_ready.emit(product)
            except Exception as e:
                self.error.emit(f"스크래핑 실패 [{url}]: {e}")
        self.finished_all.emit(products)


class AIGenerateWorker(QThread):
    """백그라운드 AI 콘텐츠 생성 (실제 Gemini/Claude 호출)"""
    progress = pyqtSignal(str, int)
    content_ready = pyqtSignal(object)
    cost_update = pyqtSignal(float)
    error = pyqtSignal(str)

    def __init__(self, product, persona: str, directive: str):
        super().__init__()
        self.product = product
        self.persona = persona
        self.directive = directive

    def run(self):
        try:
            from affiliate_system.ai_generator import AIGenerator
            gen = AIGenerator()

            self.progress.emit("훅(Hook) 생성 중...", 10)
            hook = gen.generate_hook(self.product, self.persona, self.directive)

            self.progress.emit("본문 생성 중...", 30)
            body = gen.generate_body(self.product, hook, self.persona)

            self.progress.emit("나레이션 스크립트 생성 중...", 55)
            narrations = gen.generate_narration(self.product, body)

            self.progress.emit("해시태그 생성 중...", 75)
            hashtags = gen.generate_hashtags(self.product, body)

            self.progress.emit("번역 중...", 90)
            try:
                translated = gen.translate_to_english(body[:200])
            except Exception:
                translated = ""

            content = AIContent(
                hook_text=hook or "",
                body_text=body or "",
                translated_text=translated,
                narration_scripts=narrations or [],
                hashtags=hashtags or [],
                cost_usd=gen.get_session_cost(),
                models_used=["gemini-2.5-flash", "claude-3-haiku"])

            self.progress.emit("AI 콘텐츠 생성 완료!", 100)
            self.cost_update.emit(content.cost_usd)
            self.content_ready.emit(content)
        except Exception as e:
            self.error.emit(f"AI 생성 오류: {e}")


class RenderWorker(QThread):
    """백그라운드 비디오 렌더링"""
    progress = pyqtSignal(str, int)
    render_complete = pyqtSignal(str, dict)
    error = pyqtSignal(str)

    def __init__(self, campaign_id: str):
        super().__init__()
        self.campaign_id = campaign_id

    def run(self):
        try:
            for pct in range(0, 101, 10):
                self.progress.emit(f"렌더링 중... {pct}%", pct)
                self.msleep(200)
            output_path = str(
                RENDER_OUTPUT_DIR / f"{self.campaign_id}.mp4")
            self.render_complete.emit(
                output_path,
                {"duration": 30, "size_mb": 12.5, "md5": "abc123"})
        except Exception as e:
            self.error.emit(f"렌더링 오류: {e}")


class BatchWorker(QThread):
    """MCN 대량 생산 배치 처리"""
    progress = pyqtSignal(str, int)
    row_complete = pyqtSignal(int, str)
    batch_complete = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, rows: list):
        super().__init__()
        self.rows = rows
        self._stop = False

    def run(self):
        for i, row in enumerate(self.rows):
            if self._stop:
                break
            self.progress.emit(
                f"배치 처리 중... {i+1}/{len(self.rows)}",
                int(((i + 1) / len(self.rows)) * 100))
            self.row_complete.emit(i, "완료")
            self.msleep(500)
        self.batch_complete.emit()

    def stop(self):
        self._stop = True


# ══════════════════════════════════════════════════════════════
#  TAB 1: CFO 대시보드 + 시스템 모니터
# ══════════════════════════════════════════════════════════════

class DashboardTab(QWidget):
    """ROI 추적기, 시스템 상태 LED, 실시간 콘솔"""

    def __init__(self, tracker: CostTracker, console: LiveConsole):
        super().__init__()
        self.tracker = tracker
        self.console = console
        self._init_ui()

        self._timer = QTimer(self)
        self._timer.timeout.connect(self.refresh)
        self._timer.start(15_000)

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)

        # ── 상단 지표 카드 ──
        cards = QHBoxLayout()
        cards.setSpacing(16)

        self.card_today = MetricCard(
            "오늘의 API 과금액", "₩0", "약 $0.0000", "#6366f1")
        self.card_monthly = MetricCard(
            "이번달 누적 비용", "₩0", "약 $0.0000", "#a855f7")
        self.card_campaigns = MetricCard(
            "총 캠페인", "0건", "진행중 0건", "#22c55e")
        self.card_roi = MetricCard(
            "게시물당 비용", "N/A", "데이터 없음", "#f59e0b")

        cards.addWidget(self.card_today)
        cards.addWidget(self.card_monthly)
        cards.addWidget(self.card_campaigns)
        cards.addWidget(self.card_roi)
        layout.addLayout(cards)

        # ── 예산 프로그레스 ──
        budget_frame = QFrame()
        budget_frame.setStyleSheet("""
            QFrame { background: #111827; border: 1px solid #1f2937;
                     border-radius: 12px; }
        """)
        budget_layout = QVBoxLayout(budget_frame)
        budget_layout.setContentsMargins(18, 14, 18, 14)

        bh = QHBoxLayout()
        lbl_bt = QLabel("월간 예산 현황")
        lbl_bt.setObjectName("metricLabel")
        self.lbl_budget_pct = QLabel("0%")
        self.lbl_budget_pct.setStyleSheet(
            "color: #22c55e; font-weight: 800; font-size: 15px;")
        bh.addWidget(lbl_bt)
        bh.addStretch()
        bh.addWidget(self.lbl_budget_pct)
        budget_layout.addLayout(bh)

        self.budget_bar = QProgressBar()
        self.budget_bar.setMaximum(BUDGET_LIMIT_KRW)
        self.budget_bar.setFixedHeight(12)
        budget_layout.addWidget(self.budget_bar)

        self.lbl_budget_detail = QLabel(
            f"₩0 / ₩{BUDGET_LIMIT_KRW:,}")
        self.lbl_budget_detail.setStyleSheet(
            "color: #6b7280; font-size: 11px;")
        budget_layout.addWidget(self.lbl_budget_detail)

        layout.addWidget(budget_frame)

        # ── 중간: 시스템 상태 + 캠페인 큐 ──
        mid = QHBoxLayout()
        mid.setSpacing(16)

        # 시스템 상태 LED
        status_frame = QFrame()
        status_frame.setFixedWidth(260)
        status_frame.setStyleSheet("""
            QFrame { background: #111827; border: 1px solid #1f2937;
                     border-radius: 12px; }
        """)
        status_layout = QVBoxLayout(status_frame)
        status_layout.setContentsMargins(16, 14, 16, 14)
        status_layout.setSpacing(6)

        lbl_status = QLabel("시스템 상태")
        lbl_status.setObjectName("sectionTitle")
        status_layout.addWidget(lbl_status)

        self.led_db = StatusLED("로컬 DB", True)
        # 구글 드라이브: 토큰 파일 존재 여부로 연동 상태 판단
        _drive_token = Path(__file__).parent / "workspace" / "drive_token.json"
        self.led_drive = StatusLED("구글 드라이브", _drive_token.exists())
        self.led_gemini = StatusLED("Gemini API", bool(GEMINI_API_KEY))
        self.led_claude = StatusLED("Claude API", bool(ANTHROPIC_API_KEY))
        self.led_telegram = StatusLED("텔레그램 봇", bool(TELEGRAM_BOT_TOKEN))
        self.led_chrome = StatusLED("Chrome CDP", False)
        self.led_pexels = StatusLED("Pexels (무료)", bool(PEXELS_API_KEY))
        self.led_coupang = StatusLED("쿠팡 파트너스", bool(COUPANG_ACCESS_KEY))

        status_layout.addWidget(self.led_db)
        status_layout.addWidget(self.led_drive)
        status_layout.addWidget(self.led_gemini)
        status_layout.addWidget(self.led_claude)
        status_layout.addWidget(self.led_telegram)
        status_layout.addWidget(self.led_chrome)
        status_layout.addWidget(self.led_pexels)
        status_layout.addWidget(self.led_coupang)
        status_layout.addStretch()

        # Chrome CDP 상태 체크 (비동기)
        self._check_chrome_cdp()

        mid.addWidget(status_frame)

        # 캠페인 큐 테이블
        queue_frame = QVBoxLayout()
        lbl_queue = QLabel("작업 큐")
        lbl_queue.setObjectName("sectionTitle")
        queue_frame.addWidget(lbl_queue)

        self.campaign_table = QTableWidget()
        self.campaign_table.setColumnCount(6)
        self.campaign_table.setHorizontalHeaderLabels([
            "ID", "상품/크리에이터", "상태", "플랫폼", "비용", "생성일시"])
        self.campaign_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self.campaign_table.verticalHeader().setVisible(False)
        self.campaign_table.setShowGrid(False)
        self.campaign_table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows)
        queue_frame.addWidget(self.campaign_table)

        mid.addLayout(queue_frame)
        layout.addLayout(mid)

        # ── DB 뷰어: API 사용 내역 ──
        db_frame = QFrame()
        db_frame.setStyleSheet("""
            QFrame { background: #111827; border: 1px solid #1f2937;
                     border-radius: 12px; }
        """)
        db_layout = QVBoxLayout(db_frame)
        db_layout.setContentsMargins(16, 14, 16, 14)
        db_layout.setSpacing(8)

        db_header = QHBoxLayout()
        lbl_db = QLabel("API 사용 내역 (DB)")
        lbl_db.setObjectName("sectionTitle")
        db_header.addWidget(lbl_db)
        db_header.addStretch()

        self.btn_db_refresh = QPushButton("새로고침")
        self.btn_db_refresh.setObjectName("ghostBtn")
        self.btn_db_refresh.setFixedSize(90, 32)
        self.btn_db_refresh.clicked.connect(self._load_db_records)
        db_header.addWidget(self.btn_db_refresh)
        db_layout.addLayout(db_header)

        # 모델별 요약
        self.db_summary = QLabel("데이터 로딩 중...")
        self.db_summary.setStyleSheet(
            "color: #9ca3af; font-size: 12px; padding: 4px 0;")
        db_layout.addWidget(self.db_summary)

        # 상세 기록 테이블
        self.db_table = QTableWidget()
        self.db_table.setColumnCount(6)
        self.db_table.setHorizontalHeaderLabels([
            "시각", "프로젝트", "모델", "Input 토큰",
            "Output 토큰", "비용 (USD)"])
        self.db_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch)
        self.db_table.verticalHeader().setVisible(False)
        self.db_table.setShowGrid(False)
        self.db_table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows)
        self.db_table.setMaximumHeight(200)
        self.db_table.setAlternatingRowColors(True)
        self.db_table.setStyleSheet("""
            QTableWidget { alternate-background-color: #0f1629; }
        """)
        db_layout.addWidget(self.db_table)
        layout.addWidget(db_frame)

        # ── 실시간 콘솔 ──
        lbl_console = QLabel("실시간 로그")
        lbl_console.setObjectName("sectionTitle")
        layout.addWidget(lbl_console)
        layout.addWidget(self.console)

        self.refresh()
        self._load_db_records()

    @pyqtSlot()
    def refresh(self):
        try:
            rate = self.tracker.get_exchange_rate()
            today = self.tracker.get_today_total()
            monthly = self.tracker.get_monthly_total()
            today_krw = today * rate
            monthly_krw = monthly * rate

            self.card_today.set_value(
                f"₩{today_krw:,.0f}", f"약 ${today:.4f}")
            self.card_monthly.set_value(
                f"₩{monthly_krw:,.0f}", f"약 ${monthly:.4f}")

            pct = min(int((monthly_krw / BUDGET_LIMIT_KRW) * 100), 100) \
                if BUDGET_LIMIT_KRW > 0 else 0
            self.budget_bar.setValue(int(monthly_krw))
            self.lbl_budget_pct.setText(f"{pct}%")
            self.lbl_budget_detail.setText(
                f"₩{monthly_krw:,.0f} / ₩{BUDGET_LIMIT_KRW:,}")

            color = "#ef4444" if pct >= 100 else (
                "#f59e0b" if pct >= 80 else "#22c55e")
            self.lbl_budget_pct.setStyleSheet(
                f"color: {color}; font-weight: 800; font-size: 15px;")

            # LED 상태 업데이트
            self.led_gemini.set_status(bool(GEMINI_API_KEY))
            self.led_claude.set_status(bool(ANTHROPIC_API_KEY))
            self.led_telegram.set_status(bool(TELEGRAM_BOT_TOKEN))
            self.led_pexels.set_status(bool(PEXELS_API_KEY))
            self.led_coupang.set_status(bool(COUPANG_ACCESS_KEY))
            self._check_chrome_cdp()

            # DB 뷰어 갱신
            self._load_db_records()
        except Exception:
            pass

    def _check_chrome_cdp(self):
        """Chrome CDP 연결 상태 확인"""
        try:
            import requests
            resp = requests.get(f"{CDP_URL}/json/version", timeout=2)
            self.led_chrome.set_status(resp.status_code == 200)
        except Exception:
            self.led_chrome.set_status(False)

    def _load_db_records(self):
        """api_usage.db에서 최근 기록 로드 + 비용 상세 분석"""
        try:
            import sqlite3
            from affiliate_system.config import COST_TRACKER_DB
            conn = sqlite3.connect(COST_TRACKER_DB)
            cur = conn.cursor()
            rate = self.tracker.get_exchange_rate()

            # 오늘 비용 (모델별)
            cur.execute("""
                SELECT model, COUNT(*) as calls,
                       SUM(input_tokens), SUM(output_tokens),
                       SUM(cost_usd)
                FROM api_usage
                WHERE DATE(timestamp) = DATE('now', 'localtime')
                GROUP BY model ORDER BY SUM(cost_usd) DESC
            """)
            today_rows = cur.fetchall()

            # 이번달 비용 (모델별)
            cur.execute("""
                SELECT model, COUNT(*) as calls,
                       SUM(input_tokens), SUM(output_tokens),
                       SUM(cost_usd)
                FROM api_usage
                WHERE strftime('%%Y-%%m', timestamp) =
                      strftime('%%Y-%%m', 'now', 'localtime')
                GROUP BY model ORDER BY SUM(cost_usd) DESC
            """)
            month_rows = cur.fetchall()

            # 전체 누적
            cur.execute("""
                SELECT COUNT(*), SUM(input_tokens), SUM(output_tokens),
                       SUM(cost_usd)
                FROM api_usage
            """)
            total_row = cur.fetchone()

            # 요약 텍스트 생성
            lines = []
            if today_rows:
                today_total = sum(r[4] for r in today_rows)
                today_calls = sum(r[1] for r in today_rows)
                lines.append(
                    f"[오늘] {today_calls}회 호출, "
                    f"${today_total:.4f} (₩{today_total*rate:,.0f})")
                for model, calls, t_in, t_out, cost in today_rows:
                    short = model.split("/")[-1] if "/" in model else model
                    lines.append(
                        f"  {short}: {calls}회, "
                        f"{t_in:,}+{t_out:,}tok, "
                        f"₩{cost*rate:,.0f}")

            if month_rows:
                month_total = sum(r[4] for r in month_rows)
                month_calls = sum(r[1] for r in month_rows)
                lines.append(
                    f"[이번달] {month_calls}회 호출, "
                    f"${month_total:.4f} (₩{month_total*rate:,.0f})")

            if total_row and total_row[0]:
                lines.append(
                    f"[누적] {total_row[0]:,}회, "
                    f"{total_row[1]:,}+{total_row[2]:,}tok, "
                    f"${total_row[3]:.4f} (₩{total_row[3]*rate:,.0f})")

            self.db_summary.setText(
                " | ".join(lines) if lines else "아직 API 사용 기록이 없습니다")

            # 최근 50건 상세 기록
            cur.execute("""
                SELECT timestamp, project, model,
                       input_tokens, output_tokens, cost_usd
                FROM api_usage ORDER BY id DESC LIMIT 50
            """)
            records = cur.fetchall()
            conn.close()

            self.db_table.setRowCount(len(records))
            for i, (ts, proj, model, in_tok, out_tok, cost) in enumerate(records):
                # 시각 (시:분:초만)
                short_ts = ts.split(" ")[-1] if " " in ts else ts
                self.db_table.setItem(i, 0, QTableWidgetItem(short_ts))
                self.db_table.setItem(i, 1, QTableWidgetItem(proj))
                self.db_table.setItem(i, 2, QTableWidgetItem(model))

                item_in = QTableWidgetItem(f"{in_tok:,}")
                item_in.setTextAlignment(
                    Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.db_table.setItem(i, 3, item_in)

                item_out = QTableWidgetItem(f"{out_tok:,}")
                item_out.setTextAlignment(
                    Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.db_table.setItem(i, 4, item_out)

                item_cost = QTableWidgetItem(f"${cost:.6f}")
                item_cost.setTextAlignment(
                    Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if cost > 0.01:
                    item_cost.setForeground(QColor("#f59e0b"))
                elif cost > 0.001:
                    item_cost.setForeground(QColor("#22c55e"))
                else:
                    item_cost.setForeground(QColor("#6b7280"))
                self.db_table.setItem(i, 5, item_cost)

        except Exception as e:
            self.db_summary.setText(f"DB 로드 오류: {e}")

    def update_campaigns(self, campaigns: list):
        self.campaign_table.setRowCount(len(campaigns))
        active = 0
        total_cost = 0.0
        for i, c in enumerate(campaigns):
            self.campaign_table.setItem(
                i, 0, QTableWidgetItem(c.id[:8]))
            self.campaign_table.setItem(
                i, 1, QTableWidgetItem(
                    c.product.title[:40] or c.product.url[:40]))
            bg, fg, kr = STATUS_COLORS.get(
                c.status, ("#1f2937", "#9ca3af", "대기"))
            item = QTableWidgetItem(kr)
            item.setForeground(QColor(fg))
            self.campaign_table.setItem(i, 2, item)
            plats = ", ".join(p.value for p in c.target_platforms)
            self.campaign_table.setItem(i, 3, QTableWidgetItem(plats))
            self.campaign_table.setItem(
                i, 4, QTableWidgetItem(f"${c.total_cost_usd:.4f}"))
            self.campaign_table.setItem(
                i, 5, QTableWidgetItem(
                    c.created_at.strftime("%m/%d %H:%M")))
            if c.status not in (CampaignStatus.COMPLETE,
                                CampaignStatus.ERROR,
                                CampaignStatus.DRAFT):
                active += 1
            total_cost += c.total_cost_usd

        self.card_campaigns.set_value(
            f"{len(campaigns)}건", f"진행중 {active}건")
        if len(campaigns) > 0:
            cpp = total_cost / len(campaigns)
            try:
                rate = self.tracker.get_exchange_rate()
                self.card_roi.set_value(
                    f"₩{cpp * rate:,.1f}", f"약 ${cpp:.4f}/건")
            except Exception:
                pass


# ══════════════════════════════════════════════════════════════
#  TAB 2: 3-모드 커맨드센터
# ══════════════════════════════════════════════════════════════

class ModeAWidget(QWidget):
    """Mode A: 외부 상품 판매 (제휴 마케팅) — 플랫폼 선택 + 미리보기 + 프리셋"""

    campaign_created = pyqtSignal(object)

    # 판매 최적화 페르소나 프리셋
    PERSONA_PRESETS = [
        ("직접 입력", ""),
        ("🔥 20대 여성 뷰티 유튜버", "20대 여성 뷰티 유튜버, 친근한 말투, 이모티콘 자주 사용, '언니' 호칭"),
        ("💪 30대 남성 테크 리뷰어", "30대 남성 IT/테크 리뷰어, 전문적이면서 쉬운 설명, 데이터 기반 비교"),
        ("🛍️ 쿠팡 쇼핑 전문가", "쿠팡 최저가 전문가, 할인율/로켓배송 강조, '이 가격 실화?' 식 표현"),
        ("🍳 맛집/음식 인플루언서", "맛집 탐방 인플루언서, 감탄사 많이, 먹방 느낌, 생생한 맛 표현"),
        ("👔 비즈니스 컨설턴트", "전문적 B2B 컨설턴트, 데이터와 사례 중심, 신뢰감 있는 어조"),
        ("🎮 MZ세대 트렌드세터", "MZ세대 트렌드세터, 밈/신조어 활용, 짧고 임팩트 있는 표현"),
        ("👩‍👧 육아맘 추천러", "30대 육아맘, 가성비+실용성 강조, 아이와 함께 쓸 수 있는지 중심"),
    ]

    # 판매 최적화 훅 프리셋
    HOOK_PRESETS = [
        ("직접 입력", ""),
        ("⚡ 충격 질문형", "충격적인 질문으로 시작, '이거 아직도 모르세요?' 3초 안에 클릭 유도"),
        ("💰 가격 충격형", "파격 할인가로 시작, '이 가격 실화?!' 즉각적 구매 욕구 자극"),
        ("📊 비교형", "경쟁 제품 대비 장점 비교, '○○보다 ○○이 더 좋은 이유 3가지'"),
        ("🔥 긴급/한정형", "한정 수량/시간 강조, FOMO(놓칠까봐 두려움) 자극, 카운트다운 느낌"),
        ("📖 스토리텔링형", "개인 경험담으로 시작, '나도 처음엔 몰랐는데...' 공감 유도"),
        ("🏆 베스트셀러형", "판매량/리뷰 수 강조, '100만개 팔린 이유' 사회적 증거 활용"),
        ("🎯 문제해결형", "타겟의 페인포인트 공략, '○○ 때문에 고민이라면 이것!' 해결책 제시"),
    ]

    # 소스 플랫폼 (클릭 선택)
    SOURCE_PLATFORMS = [
        ("쿠팡", "🛒", "#e44d26"),
        ("네이버", "🟢", "#03c75a"),
        ("알리", "🌏", "#ff6a00"),
        ("11번가", "🔴", "#ff0038"),
        ("티몬", "🟡", "#ff5a5f"),
        ("위메프", "🟠", "#ff4081"),
    ]

    def __init__(self):
        super().__init__()
        self._scraped_product = None
        self._init_ui()

    def _init_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(16)

        # 좌측: 입력 폼
        left = QVBoxLayout()
        left.setSpacing(10)

        # ── 소스 플랫폼 선택 + URL 입력 ──
        url_group = QGroupBox("상품 URL 입력")
        url_layout = QVBoxLayout(url_group)
        url_layout.setSpacing(8)

        # 플랫폼 선택 버튼 행
        plat_src_row = QHBoxLayout()
        plat_src_row.setSpacing(6)
        lbl_src = QLabel("소스:")
        lbl_src.setStyleSheet(
            "font-size: 11px; font-weight: 700; color: #6b7280;")
        lbl_src.setFixedWidth(35)
        plat_src_row.addWidget(lbl_src)

        self._src_buttons = []
        for name, icon, color in self.SOURCE_PLATFORMS:
            btn = QPushButton(f"{icon} {name}")
            btn.setCheckable(True)
            btn.setFixedHeight(32)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: #111827; color: #9ca3af;
                    border: 1px solid #1f2937; border-radius: 6px;
                    padding: 4px 10px; font-size: 12px; font-weight: 600;
                }}
                QPushButton:checked {{
                    background: {color}; color: white;
                    border-color: {color};
                }}
                QPushButton:hover:!checked {{
                    background: #1f2937; color: #e5e7eb;
                }}
            """)
            btn.clicked.connect(
                lambda checked, n=name: self._on_src_platform_click(n))
            plat_src_row.addWidget(btn)
            self._src_buttons.append((name, btn))
        plat_src_row.addStretch()
        url_layout.addLayout(plat_src_row)

        # URL 입력 행 (줄인 높이)
        url_row = QHBoxLayout()
        url_row.setSpacing(8)
        self.url_input = QLineEdit()
        self.url_input.setPlaceholderText(
            "상품 URL을 붙여넣기 하세요 (모든 쇼핑몰 주소 지원)")
        self.url_input.setFixedHeight(36)
        self.url_input.returnPressed.connect(self._on_scrape)
        self.btn_scrape = QPushButton("수집")
        self.btn_scrape.setObjectName("ghostBtn")
        self.btn_scrape.setFixedSize(70, 36)
        self.btn_scrape.clicked.connect(self._on_scrape)
        url_row.addWidget(self.url_input)
        url_row.addWidget(self.btn_scrape)
        url_layout.addLayout(url_row)
        left.addWidget(url_group)

        # ── 페르소나 / 훅 지시 (프리셋 포함) ──
        ai_group = QGroupBox("페르소나 & 훅 (판매 최적화)")
        ai_layout = QVBoxLayout(ai_group)
        ai_layout.setSpacing(6)

        # 페르소나 프리셋
        p_row = QHBoxLayout()
        lbl_p = QLabel("페르소나")
        lbl_p.setStyleSheet(
            "font-size: 11px; font-weight: 700; color: #6b7280;")
        lbl_p.setFixedWidth(55)
        p_row.addWidget(lbl_p)
        self.persona_combo = QComboBox()
        self.persona_combo.setFixedHeight(32)
        for label, _ in self.PERSONA_PRESETS:
            self.persona_combo.addItem(label)
        self.persona_combo.currentIndexChanged.connect(
            self._on_persona_preset)
        p_row.addWidget(self.persona_combo)
        ai_layout.addLayout(p_row)

        self.persona_input = QTextEdit()
        self.persona_input.setPlaceholderText(
            "판매 페르소나를 선택하거나 직접 입력하세요...")
        self.persona_input.setFixedHeight(44)
        ai_layout.addWidget(self.persona_input)

        # 훅 프리셋
        h_row = QHBoxLayout()
        lbl_h = QLabel("훅 전략")
        lbl_h.setStyleSheet(
            "font-size: 11px; font-weight: 700; color: #6b7280;")
        lbl_h.setFixedWidth(55)
        h_row.addWidget(lbl_h)
        self.hook_combo = QComboBox()
        self.hook_combo.setFixedHeight(32)
        for label, _ in self.HOOK_PRESETS:
            self.hook_combo.addItem(label)
        self.hook_combo.currentIndexChanged.connect(
            self._on_hook_preset)
        h_row.addWidget(self.hook_combo)
        ai_layout.addLayout(h_row)

        self.hook_input = QTextEdit()
        self.hook_input.setPlaceholderText(
            "판매 훅 전략을 선택하거나 직접 입력하세요...")
        self.hook_input.setFixedHeight(44)
        ai_layout.addWidget(self.hook_input)

        left.addWidget(ai_group)

        # ── 업로드 플랫폼 (3개 동시) ──
        plat_group = QGroupBox("업로드 플랫폼 (동시 게시)")
        plat_layout = QHBoxLayout(plat_group)
        plat_layout.setSpacing(12)
        self.chk_yt = QCheckBox("YouTube Shorts")
        self.chk_yt.setChecked(True)
        self.chk_naver = QCheckBox("네이버 블로그")
        self.chk_naver.setChecked(True)
        self.chk_ig = QCheckBox("Instagram Reels")
        self.chk_ig.setChecked(True)
        # 전체 선택/해제
        self.chk_all = QCheckBox("전체")
        self.chk_all.setChecked(True)
        self.chk_all.stateChanged.connect(self._on_check_all)
        plat_layout.addWidget(self.chk_all)
        plat_layout.addWidget(self.chk_yt)
        plat_layout.addWidget(self.chk_naver)
        plat_layout.addWidget(self.chk_ig)
        plat_layout.addStretch()

        # 자동 썸네일 옵션
        self.chk_auto_thumb = QCheckBox("자동 썸네일 생성")
        self.chk_auto_thumb.setChecked(True)
        self.chk_auto_thumb.setStyleSheet(
            "color: #a855f7; font-weight: 700;")
        plat_layout.addWidget(self.chk_auto_thumb)
        left.addWidget(plat_group)

        # ── 캠페인 생성 버튼 ──
        self.btn_generate = QPushButton("캠페인 생성 시작")
        self.btn_generate.setFixedHeight(48)
        self.btn_generate.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #6366f1, stop:1 #a855f7);
                color: white; border: none; border-radius: 12px;
                font-size: 16px; font-weight: 900;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #4f46e5, stop:1 #9333ea);
            }
        """)
        self.btn_generate.clicked.connect(self._on_generate)
        left.addWidget(self.btn_generate)
        left.addStretch()

        layout.addLayout(left, stretch=3)

        # ── 우측: 상품 미리보기 (수정됨 - 작동하게) ──
        right = QFrame()
        right.setMinimumWidth(280)
        right.setStyleSheet("""
            QFrame { background: #0a0e1a; border: 1px solid #1f2937;
                     border-radius: 14px; }
        """)
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(16, 14, 16, 14)
        right_layout.setSpacing(10)

        lbl_preview = QLabel("상품 미리보기")
        lbl_preview.setObjectName("metricLabel")
        right_layout.addWidget(lbl_preview)

        self.img_label = QLabel("상품 이미지\n미리보기")
        self.img_label.setFixedSize(250, 250)
        self.img_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.img_label.setStyleSheet("""
            background: #111827; border: 2px dashed #1f2937;
            border-radius: 14px; color: #4b5563; font-size: 14px;
        """)
        right_layout.addWidget(
            self.img_label, alignment=Qt.AlignmentFlag.AlignHCenter)

        self.lbl_title = QLabel("")
        self.lbl_title.setWordWrap(True)
        self.lbl_title.setStyleSheet(
            "font-size: 13px; font-weight: 700; color: #f9fafb;")
        self.lbl_title.setMaximumHeight(60)
        right_layout.addWidget(self.lbl_title)

        self.lbl_price = QLabel("")
        self.lbl_price.setStyleSheet(
            "font-size: 22px; font-weight: 900; color: #6366f1;")
        right_layout.addWidget(self.lbl_price)

        self.lbl_platform_tag = QLabel("")
        self.lbl_platform_tag.setStyleSheet(
            "color: #9ca3af; font-size: 11px;")
        right_layout.addWidget(self.lbl_platform_tag)

        # 썸네일 미리보기 (자동 생성용)
        self.lbl_thumb_title = QLabel("썸네일 미리보기")
        self.lbl_thumb_title.setObjectName("metricLabel")
        self.lbl_thumb_title.setVisible(False)
        right_layout.addWidget(self.lbl_thumb_title)

        self.thumb_label = QLabel("자동 썸네일로\n가야하지 않나?\n쇼츠,릴스,블로그\n셋다 가능하게")
        self.thumb_label.setFixedSize(250, 140)
        self.thumb_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.thumb_label.setStyleSheet("""
            background: #111827; border: 2px dashed #374151;
            border-radius: 10px; color: #4b5563; font-size: 12px;
        """)
        self.thumb_label.setVisible(False)
        right_layout.addWidget(
            self.thumb_label, alignment=Qt.AlignmentFlag.AlignHCenter)

        right_layout.addStretch()
        layout.addWidget(right, stretch=1)

    # ── 소스 플랫폼 클릭 ──

    def _on_src_platform_click(self, name: str):
        """소스 플랫폼 버튼 클릭 시 단일 선택 + URL 가이드"""
        for n, btn in self._src_buttons:
            btn.setChecked(n == name)
        # URL 자동 인식을 위한 플레이스홀더 업데이트
        hints = {
            "쿠팡": "https://www.coupang.com/... 상품 URL",
            "네이버": "https://smartstore.naver.com/... 상품 URL",
            "알리": "https://ko.aliexpress.com/... 상품 URL",
            "11번가": "https://www.11st.co.kr/... 상품 URL",
            "티몬": "https://www.tmon.co.kr/... 상품 URL",
            "위메프": "https://front.wemakeprice.com/... 상품 URL",
        }
        self.url_input.setPlaceholderText(hints.get(name, "상품 URL"))

    # ── 프리셋 선택 ──

    def _on_persona_preset(self, index: int):
        if index > 0:
            _, text = self.PERSONA_PRESETS[index]
            self.persona_input.setPlainText(text)

    def _on_hook_preset(self, index: int):
        if index > 0:
            _, text = self.HOOK_PRESETS[index]
            self.hook_input.setPlainText(text)

    # ── 전체 선택 ──

    def _on_check_all(self, state):
        checked = state == 2
        self.chk_yt.setChecked(checked)
        self.chk_naver.setChecked(checked)
        self.chk_ig.setChecked(checked)

    # ── 상품 수집 (개선된 스크래핑) ──

    @pyqtSlot()
    def _on_scrape(self):
        """상품 URL에서 정보를 수집한다 (멀티 플랫폼 지원)."""
        url = self.url_input.text().strip()
        if not url:
            QMessageBox.information(
                self, "입력 필요", "상품 URL을 입력해주세요.")
            return

        # URL에서 소스 플랫폼 자동 감지
        platform_detected = ""
        if "coupang.com" in url:
            platform_detected = "쿠팡"
        elif "naver.com" in url or "smartstore" in url:
            platform_detected = "네이버"
        elif "aliexpress" in url:
            platform_detected = "알리"
        elif "11st.co.kr" in url:
            platform_detected = "11번가"
        elif "tmon.co.kr" in url:
            platform_detected = "티몬"
        elif "wemakeprice" in url:
            platform_detected = "위메프"

        if platform_detected:
            for n, btn in self._src_buttons:
                btn.setChecked(n == platform_detected)
            self.lbl_platform_tag.setText(f"🏷️ {platform_detected}")

        self.btn_scrape.setEnabled(False)
        self.btn_scrape.setText("...")

        try:
            title = ""
            price = ""
            image_url = ""
            desc = ""

            # ── 쿠팡 전용 스크래퍼 사용 ──
            if "coupang.com" in url:
                try:
                    from affiliate_system.coupang_scraper import CoupangScraper
                    scraper = CoupangScraper()
                    product = scraper.scrape_product(url)
                    title = product.title or ""
                    price = product.price or ""
                    image_url = (product.image_urls[0]
                                 if product.image_urls else "")
                    desc = product.description or ""
                except Exception as ce:
                    # 쿠팡 스크래퍼 실패시 URL에서 기본 정보 추출
                    import re
                    m = re.search(r'/products/(\d+)', url)
                    prod_id = m.group(1) if m else ""
                    title = f"쿠팡 상품 #{prod_id}" if prod_id else "쿠팡 상품"
                    price = ""
                    desc = (f"쿠팡 직접 스크래핑 제한됨 (봇 차단): "
                            f"{str(ce)[:80]}\n"
                            f"→ 크롬에서 상품 페이지를 열어 수동 확인하세요")
            else:
                # ── 일반 스크래핑 (네이버/알리/기타) ──
                import requests
                from bs4 import BeautifulSoup

                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                                  'AppleWebKit/537.36 (KHTML, like Gecko) '
                                  'Chrome/131.0.0.0 Safari/537.36',
                    'Accept-Language': 'ko-KR,ko;q=0.9,en;q=0.8',
                    'Accept': 'text/html,application/xhtml+xml,'
                              'application/xml;q=0.9,*/*;q=0.8',
                    'Referer': 'https://www.google.com/',
                }
                session = requests.Session()
                resp = session.get(
                    url, headers=headers, timeout=15, allow_redirects=True)
                resp.raise_for_status()
                soup = BeautifulSoup(resp.text, 'html.parser')

                # 제목 추출
                for sel in ['meta[property="og:title"]',
                            'meta[name="title"]',
                            'h3.product_title',
                            '.topCont_headgroup__title',
                            'h1', 'title']:
                    el = soup.select_one(sel)
                    if el:
                        title = (el.get('content', '')
                                 or el.get_text(strip=True))
                        if title and len(title) > 3:
                            break

                # 가격 추출
                import re
                for sel in ['meta[property="product:price:amount"]',
                            'meta[property="og:price:amount"]',
                            'span.price', '.total_price', '#price',
                            'em.prd_price']:
                    el = soup.select_one(sel)
                    if el:
                        raw = (el.get('content', '')
                               or el.get_text(strip=True))
                        if raw:
                            nums = re.findall(r'[\d,]+', raw)
                            if nums:
                                try:
                                    price = f"₩{int(nums[0].replace(',','')):,}"
                                except ValueError:
                                    price = raw
                                break

                # 이미지 추출
                for sel in ['meta[property="og:image"]',
                            'meta[property="og:image:url"]',
                            '#repImg', '.product_thumb img']:
                    el = soup.select_one(sel)
                    if el:
                        image_url = (el.get('content', '')
                                     or el.get('src', '')
                                     or el.get('data-src', ''))
                        if image_url:
                            if image_url.startswith('//'):
                                image_url = 'https:' + image_url
                            break

                # 설명 추출
                for sel in ['meta[property="og:description"]',
                            'meta[name="description"]']:
                    el = soup.select_one(sel)
                    if el:
                        desc = el.get('content', '')
                        if desc:
                            break

            # ── UI 업데이트 ──
            if title:
                self.lbl_title.setText(title[:100])
            else:
                self.lbl_title.setText("(제목 추출 실패)")

            if price:
                self.lbl_price.setText(price)
            else:
                self.lbl_price.setText("가격 정보 없음")

            if image_url:
                try:
                    img_resp = requests.get(
                        image_url, headers=headers, timeout=10)
                    img_data = img_resp.content
                    pm = QPixmap()
                    pm.loadFromData(img_data)
                    if not pm.isNull():
                        self.img_label.setPixmap(pm.scaled(
                            250, 250,
                            Qt.AspectRatioMode.KeepAspectRatio,
                            Qt.TransformationMode.SmoothTransformation))
                        self.img_label.setStyleSheet(
                            "background: #111827; border: 1px solid #1f2937;"
                            " border-radius: 14px;")
                except Exception:
                    self.img_label.setText("이미지 로드 실패")

            # 자동 썸네일 미리보기 활성화
            if self.chk_auto_thumb.isChecked() and (title or image_url):
                self.lbl_thumb_title.setVisible(True)
                self.thumb_label.setVisible(True)
                platforms_txt = []
                if self.chk_yt.isChecked():
                    platforms_txt.append("Shorts")
                if self.chk_ig.isChecked():
                    platforms_txt.append("Reels")
                if self.chk_naver.isChecked():
                    platforms_txt.append("블로그")
                self.thumb_label.setText(
                    f"✅ 자동 썸네일 생성 예정\n"
                    f"플랫폼: {', '.join(platforms_txt)}\n"
                    f"AI가 최적 레이아웃 설계")
                self.thumb_label.setStyleSheet(
                    "background: rgba(99, 102, 241, 0.1);"
                    " border: 1px solid #6366f1;"
                    " border-radius: 10px; color: #a5b4fc;"
                    " font-size: 12px;")

            # 스크래핑된 상품 정보 저장
            self._scraped_product = {
                'title': title, 'price': price,
                'image_url': image_url, 'desc': desc,
                'url': url, 'platform': platform_detected,
            }

        except ImportError:
            QMessageBox.warning(
                self, "라이브러리 필요",
                "pip install requests beautifulsoup4\n패키지를 설치해주세요.")
        except Exception as e:
            self.lbl_title.setText(f"수집 오류: {str(e)[:60]}")
            self.lbl_price.setText("")
        finally:
            self.btn_scrape.setEnabled(True)
            self.btn_scrape.setText("수집")

    @pyqtSlot()
    def _on_generate(self):
        url = self.url_input.text().strip()
        if not url:
            QMessageBox.information(
                self, "입력 필요", "상품 URL을 입력해주세요.")
            return
        platforms = []
        if self.chk_yt.isChecked():
            platforms.append(Platform.YOUTUBE)
        if self.chk_naver.isChecked():
            platforms.append(Platform.NAVER_BLOG)
        if self.chk_ig.isChecked():
            platforms.append(Platform.INSTAGRAM)
        if not platforms:
            QMessageBox.information(
                self, "플랫폼 선택", "최소 1개 플랫폼을 선택해주세요.")
            return

        campaign = Campaign(
            id=str(uuid.uuid4())[:8],
            product=Product(url=url),
            persona=self.persona_input.toPlainText().strip(),
            hook_directive=self.hook_input.toPlainText().strip(),
            target_platforms=platforms,
            created_at=datetime.now())
        campaign.auto_thumbnail = self.chk_auto_thumb.isChecked()
        self.campaign_created.emit(campaign)
        self.url_input.clear()


class ModeBWidget(QWidget):
    """Mode B: 자사 브랜드 홍보 (F&B)"""

    campaign_created = pyqtSignal(object)

    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)

        # 브랜드 선택
        brand_group = QGroupBox("브랜드 페르소나")
        brand_layout = QHBoxLayout(brand_group)
        brand_layout.addWidget(QLabel("브랜드:"))
        self.brand_combo = QComboBox()
        self.brand_combo.addItems([
            "오레노카츠 (일식 프랜차이즈)",
            "무사짬뽕 (중화 프랜차이즈)",
            "브릿지원 (프랜차이즈 컨설팅)"])
        self.brand_combo.setFixedHeight(42)
        brand_layout.addWidget(self.brand_combo)
        layout.addWidget(brand_group)

        # 로컬 에셋 드래그 앤 드롭
        asset_group = QGroupBox("로컬 에셋 업로드")
        asset_layout = QVBoxLayout(asset_group)

        self.drop_zone = QLabel(
            "여기에 이미지/영상 파일을 드래그 앤 드롭하세요\n"
            "(PNG, JPG, MP4 지원)")
        self.drop_zone.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.drop_zone.setFixedHeight(120)
        self.drop_zone.setStyleSheet("""
            border: 2px dashed #1f2937; border-radius: 14px;
            color: #4b5563; font-size: 14px; padding: 20px;
            background: #0a0e1a;
        """)
        asset_layout.addWidget(self.drop_zone)

        self.file_list = QLabel("")
        self.file_list.setStyleSheet("color: #6b7280; font-size: 12px;")
        asset_layout.addWidget(self.file_list)
        layout.addWidget(asset_group)

        # 콘텐츠 유형 선택
        content_group = QGroupBox("콘텐츠 설정")
        content_layout = QGridLayout(content_group)
        content_layout.setSpacing(12)

        content_layout.addWidget(QLabel("콘텐츠 유형:"), 0, 0)
        self.content_combo = QComboBox()
        self.content_combo.addItems([
            "YouTube Shorts (세로형 숏폼)",
            "네이버 블로그 SEO 포스트",
            "Instagram Reels + DM 유도",
            "전체 플랫폼 동시 게시"])
        content_layout.addWidget(self.content_combo, 0, 1)

        content_layout.addWidget(QLabel("톤 & 매너:"), 1, 0)
        self.tone_combo = QComboBox()
        self.tone_combo.addItems([
            "프로페셔널 B2B 컨설팅",
            "친근한 맛집 리뷰",
            "트렌디 MZ세대 타겟",
            "정보 전달형 (교육)"])
        content_layout.addWidget(self.tone_combo, 1, 1)

        layout.addWidget(content_group)

        # 생성 버튼
        self.btn_generate = QPushButton("브랜드 콘텐츠 생성")
        self.btn_generate.setFixedHeight(50)
        self.btn_generate.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #f59e0b, stop:1 #ef4444);
                color: white; border: none; border-radius: 12px;
                font-size: 15px; font-weight: 900;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #d97706, stop:1 #dc2626);
            }
        """)
        self.btn_generate.clicked.connect(self._on_generate)
        layout.addWidget(self.btn_generate)
        layout.addStretch()

        self._dropped_files = []

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.drop_zone.setStyleSheet("""
                border: 2px solid #6366f1; border-radius: 14px;
                background: rgba(99, 102, 241, 0.08);
                color: #e5e7eb; font-size: 14px; padding: 20px;
            """)

    def dragLeaveEvent(self, event):
        self.drop_zone.setStyleSheet("""
            border: 2px dashed #1f2937; border-radius: 14px;
            color: #4b5563; font-size: 14px; padding: 20px;
            background: #0a0e1a;
        """)

    def dropEvent(self, event: QDropEvent):
        self.drop_zone.setStyleSheet("""
            border: 2px solid #22c55e; border-radius: 14px;
            color: #22c55e; font-size: 14px; padding: 20px;
            background: rgba(34, 197, 94, 0.05);
        """)
        self._dropped_files = [
            url.toLocalFile() for url in event.mimeData().urls()]
        names = [Path(f).name for f in self._dropped_files[:5]]
        self.drop_zone.setText(
            f"{len(self._dropped_files)}개 파일 로드 완료")
        self.file_list.setText("  |  ".join(names))

    @pyqtSlot()
    def _on_generate(self):
        brand_map = {0: "오레노카츠", 1: "무사짬뽕", 2: "브릿지원"}
        brand = brand_map.get(self.brand_combo.currentIndex(), "")
        campaign = Campaign(
            id=str(uuid.uuid4())[:8],
            product=Product(title=f"[{brand}] 브랜드 홍보"),
            persona=brand,
            target_platforms=[
                Platform.YOUTUBE, Platform.NAVER_BLOG, Platform.INSTAGRAM],
            created_at=datetime.now())
        self.campaign_created.emit(campaign)


class ModeCWidget(QWidget):
    """Mode C: MCN 크리에이터 대량 생산"""

    campaign_created = pyqtSignal(object)
    batch_started = pyqtSignal(list)

    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)

        # 엑셀 업로드
        xlsx_group = QGroupBox("대량 생산 Excel 업로드 (.xlsx)")
        xlsx_layout = QVBoxLayout(xlsx_group)

        self.drop_zone = QLabel(
            ".xlsx 파일을 여기에 드래그 앤 드롭하세요\n"
            "[크리에이터명, 주제, 에셋경로] 형식")
        self.drop_zone.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.drop_zone.setFixedHeight(100)
        self.drop_zone.setStyleSheet("""
            border: 2px dashed #1f2937; border-radius: 14px;
            color: #4b5563; font-size: 14px; padding: 20px;
            background: #0a0e1a;
        """)
        self.drop_zone.mousePressEvent = lambda e: self._browse_xlsx()
        xlsx_layout.addWidget(self.drop_zone)

        self.xlsx_table = QTableWidget()
        self.xlsx_table.setColumnCount(4)
        self.xlsx_table.setHorizontalHeaderLabels([
            "크리에이터", "주제", "에셋 경로", "상태"])
        self.xlsx_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self.xlsx_table.verticalHeader().setVisible(False)
        self.xlsx_table.setShowGrid(False)
        self.xlsx_table.setVisible(False)
        xlsx_layout.addWidget(self.xlsx_table)

        layout.addWidget(xlsx_group)

        # 배치 설정
        batch_group = QGroupBox("배치 설정")
        batch_layout = QGridLayout(batch_group)
        batch_layout.setSpacing(12)

        batch_layout.addWidget(QLabel("AI 엔진:"), 0, 0)
        self.ai_combo = QComboBox()
        self.ai_combo.addItems([
            "Claude 3 Sonnet (고품질 스크립트)",
            "Gemini 2.5 Flash (빠른 생성)",
            "하이브리드 (Claude + Gemini)"])
        batch_layout.addWidget(self.ai_combo, 0, 1)

        batch_layout.addWidget(QLabel("업로드 간격:"), 1, 0)
        self.delay_spin = QSpinBox()
        self.delay_spin.setRange(5, 60)
        self.delay_spin.setValue(25)
        self.delay_spin.setSuffix("분")
        batch_layout.addWidget(self.delay_spin, 1, 1)

        layout.addWidget(batch_group)

        # 진행 상황
        progress_group = QGroupBox("배치 진행 상황")
        progress_layout = QVBoxLayout(progress_group)

        self.batch_progress = QProgressBar()
        self.batch_progress.setFixedHeight(14)
        progress_layout.addWidget(self.batch_progress)

        prog_stats = QHBoxLayout()
        self.lbl_completed = QLabel("완료: 0건")
        self.lbl_completed.setStyleSheet(
            "color: #22c55e; font-weight: 700;")
        self.lbl_remaining = QLabel("남은 작업: 0건")
        self.lbl_remaining.setStyleSheet(
            "color: #f59e0b; font-weight: 700;")
        self.lbl_errors = QLabel("오류: 0건")
        self.lbl_errors.setStyleSheet(
            "color: #ef4444; font-weight: 700;")
        prog_stats.addWidget(self.lbl_completed)
        prog_stats.addWidget(self.lbl_remaining)
        prog_stats.addWidget(self.lbl_errors)
        prog_stats.addStretch()
        progress_layout.addLayout(prog_stats)

        layout.addWidget(progress_group)

        # 실행 / 중지 버튼
        btn_row = QHBoxLayout()
        self.btn_start = QPushButton("야간 대량 생산 시작")
        self.btn_start.setFixedHeight(52)
        self.btn_start.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #22c55e, stop:1 #16a34a);
                color: white; border: none; border-radius: 12px;
                font-size: 15px; font-weight: 900;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #16a34a, stop:1 #15803d);
            }
        """)
        self.btn_start.clicked.connect(self._on_start_batch)
        btn_row.addWidget(self.btn_start)

        self.btn_stop = QPushButton("중지")
        self.btn_stop.setObjectName("dangerBtn")
        self.btn_stop.setFixedSize(100, 52)
        self.btn_stop.setEnabled(False)
        btn_row.addWidget(self.btn_stop)

        layout.addLayout(btn_row)
        layout.addStretch()

        self._xlsx_rows = []

    def _browse_xlsx(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Excel 파일 선택", "",
            "Excel (*.xlsx);;모든 파일 (*)")
        if path:
            self._load_xlsx(path)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                if url.toLocalFile().endswith('.xlsx'):
                    event.acceptProposedAction()
                    return

    def dropEvent(self, event: QDropEvent):
        for url in event.mimeData().urls():
            path = url.toLocalFile()
            if path.endswith('.xlsx'):
                self._load_xlsx(path)

    def _load_xlsx(self, file_path: str):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()

            self._xlsx_rows = rows
            self.xlsx_table.setVisible(True)
            self.xlsx_table.setRowCount(len(rows))
            for i, row in enumerate(rows):
                for j in range(min(3, len(row))):
                    self.xlsx_table.setItem(
                        i, j, QTableWidgetItem(str(row[j] or "")))
                self.xlsx_table.setItem(
                    i, 3, QTableWidgetItem("대기"))

            self.drop_zone.setText(
                f"로드 완료: {Path(file_path).name} ({len(rows)}건)")
            self.drop_zone.setStyleSheet("""
                border: 2px solid #22c55e; border-radius: 14px;
                color: #22c55e; font-size: 14px; padding: 20px;
                background: rgba(34, 197, 94, 0.05);
            """)
            self.lbl_remaining.setText(f"남은 작업: {len(rows)}건")
            self.batch_progress.setMaximum(len(rows))
        except Exception as e:
            QMessageBox.warning(
                self, "파일 오류", f"엑셀 로드 실패: {e}")

    @pyqtSlot()
    def _on_start_batch(self):
        if not self._xlsx_rows:
            QMessageBox.information(
                self, "파일 필요", "먼저 .xlsx 파일을 업로드해주세요.")
            return
        self.batch_started.emit(self._xlsx_rows)


class CommandCenterTab(QWidget):
    """3-모드 커맨드센터 (메인 작업 탭)"""

    campaign_created = pyqtSignal(object)
    batch_started = pyqtSignal(list)

    def __init__(self):
        super().__init__()
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)

        # 모드 선택 버튼
        mode_row = QHBoxLayout()
        mode_row.setSpacing(12)

        self.btn_mode_a = QPushButton("Mode A: 외부 상품 제휴")
        self.btn_mode_b = QPushButton("Mode B: 자사 브랜드 홍보")
        self.btn_mode_c = QPushButton("Mode C: MCN 대량 생산")

        for btn in [self.btn_mode_a, self.btn_mode_b, self.btn_mode_c]:
            btn.setFixedHeight(46)
            btn.setCheckable(True)
            btn.setStyleSheet("""
                QPushButton {
                    background: #111827; color: #9ca3af;
                    border: 1px solid #1f2937; border-radius: 10px;
                    font-weight: 700; font-size: 13px;
                }
                QPushButton:checked {
                    background: #6366f1; color: white;
                    border-color: #6366f1;
                }
                QPushButton:hover:!checked {
                    background: #1f2937; color: #e5e7eb;
                }
            """)

        self.btn_mode_a.setChecked(True)
        mode_row.addWidget(self.btn_mode_a)
        mode_row.addWidget(self.btn_mode_b)
        mode_row.addWidget(self.btn_mode_c)
        layout.addLayout(mode_row)

        # 스택 위젯 (모드별 화면)
        self.stack = QStackedWidget()
        self.mode_a = ModeAWidget()
        self.mode_b = ModeBWidget()
        self.mode_c = ModeCWidget()

        self.stack.addWidget(self.mode_a)
        self.stack.addWidget(self.mode_b)
        self.stack.addWidget(self.mode_c)
        layout.addWidget(self.stack)

        # 시그널 연결
        self.btn_mode_a.clicked.connect(lambda: self._switch_mode(0))
        self.btn_mode_b.clicked.connect(lambda: self._switch_mode(1))
        self.btn_mode_c.clicked.connect(lambda: self._switch_mode(2))

        self.mode_a.campaign_created.connect(self.campaign_created.emit)
        self.mode_b.campaign_created.connect(self.campaign_created.emit)
        self.mode_c.campaign_created.connect(self.campaign_created.emit)
        self.mode_c.batch_started.connect(self.batch_started.emit)

    def _switch_mode(self, idx: int):
        self.stack.setCurrentIndex(idx)
        for i, btn in enumerate(
                [self.btn_mode_a, self.btn_mode_b, self.btn_mode_c]):
            btn.setChecked(i == idx)


# ══════════════════════════════════════════════════════════════
#  TAB 3: 설정
# ══════════════════════════════════════════════════════════════

class SettingsTab(QWidget):
    """API 키, AI 모델 라우팅, 예산 설정"""

    settings_saved = pyqtSignal()

    def __init__(self):
        super().__init__()
        self._init_ui()

    def _init_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        container = QWidget()
        inner = QVBoxLayout(container)
        inner.setContentsMargins(20, 20, 20, 20)
        inner.setSpacing(16)

        # API 키
        keys_group = QGroupBox("API 키 관리")
        keys_layout = QGridLayout(keys_group)
        keys_layout.setSpacing(12)

        keys_layout.addWidget(QLabel("Gemini API:"), 0, 0)
        self.gemini_key = QLineEdit()
        self.gemini_key.setEchoMode(QLineEdit.EchoMode.Password)
        self.gemini_key.setText(GEMINI_API_KEY)
        self.gemini_key.setPlaceholderText("AIzaSy...")
        keys_layout.addWidget(self.gemini_key, 0, 1)

        keys_layout.addWidget(QLabel("Claude API:"), 1, 0)
        self.claude_key = QLineEdit()
        self.claude_key.setEchoMode(QLineEdit.EchoMode.Password)
        self.claude_key.setText(ANTHROPIC_API_KEY)
        self.claude_key.setPlaceholderText("sk-ant-...")
        keys_layout.addWidget(self.claude_key, 1, 1)

        inner.addWidget(keys_group)

        # 스톡 미디어 API 키
        stock_group = QGroupBox("스톡 미디어 API 키 (무료 이미지/영상 검색)")
        stock_layout = QGridLayout(stock_group)
        stock_layout.setSpacing(12)

        stock_layout.addWidget(QLabel("Pexels API:"), 0, 0)
        self.pexels_key = QLineEdit()
        self.pexels_key.setEchoMode(QLineEdit.EchoMode.Password)
        self.pexels_key.setText(PEXELS_API_KEY)
        self.pexels_key.setPlaceholderText("pexels.com에서 무료 발급")
        stock_layout.addWidget(self.pexels_key, 0, 1)

        stock_layout.addWidget(QLabel("Pixabay API:"), 1, 0)
        self.pixabay_key = QLineEdit()
        self.pixabay_key.setEchoMode(QLineEdit.EchoMode.Password)
        self.pixabay_key.setText(PIXABAY_API_KEY)
        self.pixabay_key.setPlaceholderText("pixabay.com에서 무료 발급")
        stock_layout.addWidget(self.pixabay_key, 1, 1)

        stock_layout.addWidget(QLabel("Unsplash API:"), 2, 0)
        self.unsplash_key = QLineEdit()
        self.unsplash_key.setEchoMode(QLineEdit.EchoMode.Password)
        self.unsplash_key.setText(UNSPLASH_ACCESS_KEY)
        self.unsplash_key.setPlaceholderText("unsplash.com에서 무료 발급")
        stock_layout.addWidget(self.unsplash_key, 2, 1)

        inner.addWidget(stock_group)

        # AI 모델 라우팅 (NO OpenAI)
        model_group = QGroupBox("AI 모델 라우팅 (비용 최적화)")
        model_layout = QGridLayout(model_group)
        model_layout.setSpacing(12)

        model_layout.addWidget(QLabel("스크립트 생성:"), 0, 0)
        self.script_model = QComboBox()
        self.script_model.addItems([
            "Claude 3 Sonnet (고품질 추천)",
            "Claude 3 Haiku (저비용)",
            "Gemini 2.5 Flash (최저가)"])
        model_layout.addWidget(self.script_model, 0, 1)

        model_layout.addWidget(QLabel("이미지/영상 생성:"), 1, 0)
        self.visual_model = QComboBox()
        self.visual_model.addItems([
            "Gemini Imagen (이미지 생성)",
            "Gemini Veo (B-roll 영상 생성)"])
        model_layout.addWidget(self.visual_model, 1, 1)

        model_layout.addWidget(QLabel("번역/구조화:"), 2, 0)
        self.trans_model = QComboBox()
        self.trans_model.addItems([
            "Claude 3 Haiku (추천)",
            "Gemini 2.5 Flash"])
        model_layout.addWidget(self.trans_model, 2, 1)

        inner.addWidget(model_group)

        # 예산
        budget_group = QGroupBox("예산 관리")
        budget_layout = QGridLayout(budget_group)
        budget_layout.setSpacing(12)

        budget_layout.addWidget(QLabel("월간 한도:"), 0, 0)
        self.budget_spin = QSpinBox()
        self.budget_spin.setRange(1_000, 500_000)
        self.budget_spin.setValue(BUDGET_LIMIT_KRW)
        self.budget_spin.setSuffix(" 원")
        self.budget_spin.setSingleStep(5_000)
        budget_layout.addWidget(self.budget_spin, 0, 1)

        budget_layout.addWidget(QLabel("경고 임계값:"), 1, 0)
        self.warn_spin = QSpinBox()
        self.warn_spin.setRange(50, 100)
        self.warn_spin.setValue(80)
        self.warn_spin.setSuffix("%")
        budget_layout.addWidget(self.warn_spin, 1, 1)

        inner.addWidget(budget_group)

        # Instagram
        ig_group = QGroupBox("Instagram 계정")
        ig_layout = QGridLayout(ig_group)
        ig_layout.addWidget(QLabel("사용자명:"), 0, 0)
        self.ig_user = QLineEdit()
        self.ig_user.setText(INSTAGRAM_USERNAME)
        ig_layout.addWidget(self.ig_user, 0, 1)
        ig_layout.addWidget(QLabel("비밀번호:"), 1, 0)
        self.ig_pass = QLineEdit()
        self.ig_pass.setEchoMode(QLineEdit.EchoMode.Password)
        ig_layout.addWidget(self.ig_pass, 1, 1)
        inner.addWidget(ig_group)

        # 네이버
        naver_group = QGroupBox("네이버 블로그")
        naver_layout = QGridLayout(naver_group)
        naver_layout.addWidget(QLabel("Blog ID:"), 0, 0)
        self.naver_id = QLineEdit()
        self.naver_id.setText("jyjzzj")
        naver_layout.addWidget(self.naver_id, 0, 1)
        naver_layout.addWidget(QLabel("Chrome CDP:"), 1, 0)
        self.cdp_url = QLineEdit()
        self.cdp_url.setText("http://127.0.0.1:9222")
        naver_layout.addWidget(self.cdp_url, 1, 1)
        inner.addWidget(naver_group)

        # 저장 버튼
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.btn_save = QPushButton("설정 저장")
        self.btn_save.setObjectName("successBtn")
        self.btn_save.setFixedHeight(48)
        self.btn_save.setFixedWidth(200)
        self.btn_save.clicked.connect(self._save)
        btn_row.addWidget(self.btn_save)
        inner.addLayout(btn_row)
        inner.addStretch()

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def _save(self):
        env_path = PROJECT_DIR / '.env'
        existing = {}
        if env_path.exists():
            with open(env_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if '=' in line and not line.startswith('#'):
                        k, v = line.split('=', 1)
                        existing[k.strip()] = v.strip()

        existing['GEMINI_API_KEY'] = self.gemini_key.text().strip()
        existing['ANTHROPIC_API_KEY'] = self.claude_key.text().strip()
        existing['INSTAGRAM_USERNAME'] = self.ig_user.text().strip()
        existing['PEXELS_API_KEY'] = self.pexels_key.text().strip()
        existing['PIXABAY_API_KEY'] = self.pixabay_key.text().strip()
        existing['UNSPLASH_ACCESS_KEY'] = self.unsplash_key.text().strip()

        with open(env_path, 'w', encoding='utf-8') as f:
            for k, v in existing.items():
                f.write(f"{k}={v}\n")

        self.settings_saved.emit()
        QMessageBox.information(self, "저장 완료", "설정이 .env에 저장되었습니다.")


# ══════════════════════════════════════════════════════════════
#  메인 윈도우
# ══════════════════════════════════════════════════════════════

class MainWindow(QMainWindow):
    """YJ Partners MCN & F&B 자동화 — 커맨드센터"""

    def __init__(self):
        super().__init__()
        self.setWindowTitle(
            "YJ Partners MCN & F&B 자동화 파이프라인 — BRIDGE ONE")
        self.setMinimumSize(1300, 860)
        self.resize(1440, 920)

        # 핵심 서비스
        self.tracker = CostTracker(
            db_path=COST_TRACKER_DB,
            project_name="affiliate_system")

        self.campaigns: list[Campaign] = []
        self.console = LiveConsole()

        self._batch_worker = None

        self._init_ui()
        self._connect_signals()

        # 상태바 자동 갱신
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._update_status_bar)
        self._timer.start(10_000)
        self._update_status_bar()

        self.console.log("커맨드센터 시작 완료")
        self.console.log(
            f"Gemini API: {'연결됨' if GEMINI_API_KEY else '미설정'}")
        self.console.log(
            f"Claude API: {'연결됨' if ANTHROPIC_API_KEY else '미설정'}")

    def _init_ui(self):
        self.setStyleSheet(DARK_STYLESHEET)

        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)

        self.dashboard_tab = DashboardTab(self.tracker, self.console)
        self.command_tab = CommandCenterTab()
        self.editor_tab = EditorTab()
        self.db_viewer_tab = DBViewerTab()
        self.ai_review_tab = AIReviewTab()
        self.settings_tab = SettingsTab()

        self.tabs.addTab(self.dashboard_tab,  "  대시보드  ")
        self.tabs.addTab(self.command_tab,    "  작업 센터  ")
        self.tabs.addTab(self.editor_tab,     "  편집  ")
        self.tabs.addTab(self.db_viewer_tab,  "  DB 뷰어  ")
        self.tabs.addTab(self.ai_review_tab,  "  AI 검토  ")
        self.tabs.addTab(self.settings_tab,   "  설정  ")

        self.setCentralWidget(self.tabs)

        # 상태바
        sb = QStatusBar()
        self.setStatusBar(sb)

        self.cost_label = QLabel("  오늘 비용: ₩0  ")
        self.cost_label.setStyleSheet(
            "color: #9ca3af; font-weight: 700; font-size: 12px;")
        sb.addPermanentWidget(self.cost_label)

        self.campaigns_label = QLabel("  캠페인: 0건  ")
        self.campaigns_label.setStyleSheet(
            "color: #6b7280; font-size: 12px;")
        sb.addWidget(self.campaigns_label)

    def _connect_signals(self):
        self.command_tab.campaign_created.connect(
            self._on_campaign_created)
        self.command_tab.batch_started.connect(
            self._on_batch_started)
        self.settings_tab.settings_saved.connect(
            lambda: self.console.log("설정 저장 완료"))

        # 편집 탭 → AI 검토 연동
        self.editor_tab.send_to_review.connect(
            self._on_editor_to_review)
        # 편집 탭 → 구글 드라이브 업로드
        self.editor_tab.upload_to_drive.connect(
            self._on_editor_drive_upload)

    @pyqtSlot(object)
    def _on_campaign_created(self, campaign: Campaign):
        self.campaigns.append(campaign)
        campaign.status = CampaignStatus.DRAFT
        self.dashboard_tab.update_campaigns(self.campaigns)
        self.campaigns_label.setText(
            f"  캠페인: {len(self.campaigns)}건  ")
        self.console.log(
            f"캠페인 생성: {campaign.id} → "
            f"{', '.join(p.value for p in campaign.target_platforms)}")

        # 편집 탭에 캠페인 데이터 전달 (작업센터 → 편집 연동)
        scraped = getattr(
            self.command_tab.mode_a, '_scraped_product', None)
        campaign_data = {
            'id': campaign.id,
            'title': (scraped or {}).get('title', ''),
            'url': campaign.product.url,
            'image_url': (scraped or {}).get('image_url', ''),
            'platforms': [p.value for p in campaign.target_platforms],
            'persona': campaign.persona,
            'hook': campaign.hook_directive,
        }
        self.editor_tab.load_campaign(campaign_data)
        self.tabs.setCurrentWidget(self.editor_tab)
        self.console.log("편집 탭으로 캠페인 데이터 전달 완료")

    @pyqtSlot(dict)
    def _on_editor_to_review(self, review_data: dict):
        """편집 탭 → AI 검토 탭 연동"""
        self.console.log(
            f"AI 검토 요청: 마커 {len(review_data.get('markers', []))}개, "
            f"플랫폼: {review_data.get('platform', '전체')}")
        self.tabs.setCurrentWidget(self.ai_review_tab)

    @pyqtSlot(dict)
    def _on_editor_drive_upload(self, upload_data: dict):
        """편집 탭 → Google Drive 업로드"""
        self.console.log(
            f"Google Drive 업로드: {len(upload_data.get('files', []))}개 파일, "
            f"플랫폼: {upload_data.get('platform', 'unknown')}")

    @pyqtSlot(list)
    def _on_batch_started(self, rows: list):
        self.console.log(f"MCN 대량 생산 시작: {len(rows)}건")
        mode_c = self.command_tab.mode_c
        mode_c.btn_start.setEnabled(False)
        mode_c.btn_stop.setEnabled(True)

        self._batch_worker = BatchWorker(rows)
        self._batch_worker.progress.connect(
            lambda msg, pct: (
                mode_c.batch_progress.setValue(pct),
                self.console.log(msg)))
        self._batch_worker.row_complete.connect(
            lambda i, s: (
                mode_c.xlsx_table.setItem(i, 3, QTableWidgetItem(s)),
                mode_c.lbl_completed.setText(f"완료: {i+1}건"),
                mode_c.lbl_remaining.setText(
                    f"남은 작업: {len(rows) - i - 1}건")))
        self._batch_worker.batch_complete.connect(
            lambda: (
                mode_c.btn_start.setEnabled(True),
                mode_c.btn_stop.setEnabled(False),
                self.console.log("배치 처리 완료!")))
        self._batch_worker.start()

        mode_c.btn_stop.clicked.connect(self._batch_worker.stop)

    def _update_status_bar(self):
        try:
            today = self.tracker.get_today_total()
            rate = self.tracker.get_exchange_rate()
            self.cost_label.setText(
                f"  오늘 비용: ₩{today * rate:,.0f}  ")
        except Exception:
            pass

    def closeEvent(self, event):
        reply = QMessageBox.question(
            self, "종료 확인",
            "커맨드센터를 종료하시겠습니까?",
            QMessageBox.StandardButton.Yes |
            QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            if self._batch_worker and self._batch_worker.isRunning():
                self._batch_worker.stop()
                self._batch_worker.wait(3000)
            event.accept()
        else:
            event.ignore()


# ══════════════════════════════════════════════════════════════
#  엔트리 포인트
# ══════════════════════════════════════════════════════════════

def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    palette = app.palette()
    palette.setColor(palette.ColorRole.Window, QColor("#0a0e1a"))
    palette.setColor(palette.ColorRole.WindowText, QColor("#e5e7eb"))
    palette.setColor(palette.ColorRole.Base, QColor("#111827"))
    palette.setColor(palette.ColorRole.AlternateBase, QColor("#0a0e1a"))
    palette.setColor(palette.ColorRole.Text, QColor("#e5e7eb"))
    palette.setColor(palette.ColorRole.Button, QColor("#1f2937"))
    palette.setColor(palette.ColorRole.ButtonText, QColor("#e5e7eb"))
    palette.setColor(palette.ColorRole.Highlight, QColor("#6366f1"))
    palette.setColor(palette.ColorRole.HighlightedText, QColor("#ffffff"))
    app.setPalette(palette)

    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
